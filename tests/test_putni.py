# -*- coding: utf-8 -*-
"""Testovi čitanja putnih naloga iz mjesečnog PN Excela (jedan list = jedan nalog)."""
from datetime import datetime

import openpyxl

from src import putni as P


def _napravi_pn_excel(path):
    """Izmišljeni PN Excel s 3 lista: normalan nalog, nalog s 0 € i prazan list."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "071"
    ws["A1"] = "Putni nalog broj:"; ws["B1"] = "071"
    ws["A2"] = "Djelatnik:";        ws["B2"] = "Ivan Horvat"
    ws["A3"] = "Datum odlaska:";    ws["B3"] = datetime(2026, 4, 14)
    ws["A4"] = "Za isplatu:";       ws["B4"] = 0        # kratki put BEZ dnevnice -> MORA ući!

    ws2 = wb.create_sheet("072")
    ws2["A1"] = "Putni nalog broj:"; ws2["B1"] = "072"
    ws2["A2"] = "Djelatnik:";        ws2["B2"] = "Ana Anić"
    ws2["A3"] = "Datum odlaska:";    ws2["B3"] = datetime(2026, 4, 15)
    ws2["A4"] = "Za isplatu:";       ws2["B4"] = 30.5

    ws3 = wb.create_sheet("prazan")   # bez djelatnika/iznosa -> preskače se
    ws3["A1"] = "Putni nalog broj:"

    wb.save(path)


def test_procitaj_naloge(tmp_path):
    p = str(tmp_path / "pn 04-2026.xlsx")
    _napravi_pn_excel(p)
    nalozi = P.procitaj_naloge(p)

    brojevi = sorted(n["broj"] for n in nalozi)
    assert brojevi == ["071", "072"]          # prazan list preskočen

    n071 = next(n for n in nalozi if n["broj"] == "071")
    assert n071["iznos"] == 0                 # 0 € je VALJAN nalog (ne preskače se)
    assert n071["djelatnik"] == "Ivan Horvat"
    assert n071["datum"] == datetime(2026, 4, 14).date()

    n072 = next(n for n in nalozi if n["broj"] == "072")
    assert n072["iznos"] == 30.5


def test_broj_se_normalizira(tmp_path):
    # broj '33' u listu -> '033' (troznamenkasto, kao u knjizi)
    p = str(tmp_path / "pn 01-2026.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Putni nalog broj:"; ws["B1"] = "33"
    ws["A2"] = "Djelatnik:";        ws["B2"] = "Pero Perić"
    ws["A3"] = "Datum odlaska:";    ws["B3"] = datetime(2026, 1, 5)
    ws["A4"] = "Za isplatu:";       ws["B4"] = 15
    wb.save(p)

    nalozi = P.procitaj_naloge(p)
    assert nalozi[0]["broj"] == "033"
