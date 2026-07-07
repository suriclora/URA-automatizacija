# -*- coding: utf-8 -*-
"""
URA — automatizacija obrade ulaznih računa (URA).

Pokretanje: dvostruki klik na 'pokreni.bat' (ili: python main.py)

Na početku biraš:
  1) SAMO PRIKAŽI — ništa ne dira, samo kaže što je novo
  2) ODRADI SVE   — upiše nove račune + razvrsta i poveže datoteke + sve ostalo

Tijek datoteka (opcija 2, pravi rad):
  - XML i Parrin PDF  -> Računi-URA/2026/Parra/{Mjesec}/
  - originalni PDF dobavljača -> Računi-URA/2026/pdf/ kao "UR XXXX.pdf" (+ link u Excel)

Sigurnosne sklopke u config.py:
  KORISTI_TEST_EXCEL = True -> radi na kopiji; datoteke se NE premještaju (samo prikaz)
  SIMULACIJA = True         -> potpuno ništa ne mijenja (kao opcija 1)
"""
import os
import shutil
import traceback
from datetime import datetime, date, timedelta

from rapidfuzz import fuzz
from openpyxl.styles import Font

import config
from src.log_setup import postavi_logging
from src import backup as backup_mod
from src import excel as excel_mod
from src import parra as parra_mod
from src import pdf_utils
from src import dedup as dedup_mod
from src import datoteke as dat_mod
from src import spajanje as spajanje_mod
from src import upis_izvoda as izvod_mod
from src import vrsta as vrsta_mod
from src import stanje as stanje_mod
from src import putni as putni_mod
from src import izvjestaji as izvjestaji_mod
from src import racuni_ocr as ocr_mod
from src import fotke as fotke_mod
from src import putni_izrada as putni_izrada_mod
from src import vremena as vremena_mod
from src.utils import ocisti


# ---------------------------------------------------------------------
#  Male pomoćne funkcije
# ---------------------------------------------------------------------

def _parse_datum(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _broj(x):
    try:
        return float(x or 0)
    except (ValueError, TypeError):
        return 0.0


def _parse_datum_hr(s):
    """'18.05.2026' -> date; None ako ne valja."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%d.%m.%Y").date()
    except (ValueError, TypeError):
        return None


def _datum_celije(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---------------------------------------------------------------------
#  DEFERNO SPAJANJE  (uplata stigla prije računa -> UR0 red čeka; kad
#  račun stigne, dopuni taj red umjesto da radimo duplikat)
# ---------------------------------------------------------------------

def _indeks_ceka_racun(ws, obradjuj_od):
    """UR0 redovi (od datuma starta nadalje) koji ČEKAJU račun: imaju dobavljača i
    plaćeno, a NEMAJU broj računa. Kandidati za deferno spajanje s e-računom koji tek
    stiže. Čita se iz radnog lista (vrijednosti su literali, ne formule)."""
    out = []
    for i in range(config.HEADER_RED + 1, ws.max_row + 1):
        if str(ws.cell(i, 1).value).split(".")[0] != "0":      # samo UR 0000
            continue
        if ws.cell(i, 2).value not in (None, ""):              # već ima broj računa
            continue
        dob = ws.cell(i, 5).value
        plac = ws.cell(i, 16).value
        if not dob or not isinstance(plac, (int, float)):      # treba dobavljač + plaćeno
            continue
        datum = _datum_celije(ws.cell(i, 4).value)
        if obradjuj_od and datum and datum < obradjuj_od:      # ne diraj ručni zaostatak
            continue
        out.append({"row": i, "dob_clean": ocisti(dob), "iznos": float(plac),
                    "datum": datum, "iskoristen": False})
    return out


def _deferno_kandidat(dob_clean, iznos, datum_rac, ceka):
    """Vrati JEDINSTVENI UR0 red koji odgovara e-računu (isti dobavljač + isti iznos,
    a datum računa nije nakon datuma uplate). None ako nema ili ih je više (ne nagađamo)."""
    pog = []
    for c in ceka:
        if c["iskoristen"] or abs(c["iznos"] - iznos) > config.DEDUP_TOL_IZNOS:
            continue
        if not (dob_clean and c["dob_clean"]):
            continue
        if fuzz.partial_ratio(dob_clean, c["dob_clean"]) < spajanje_mod.SIM_IME:
            continue
        if datum_rac and c["datum"] and datum_rac > c["datum"] + timedelta(days=2):
            continue  # račun ne može biti (znatno) nakon uplate
        pog.append(c)
    return pog[0] if len(pog) == 1 else None


def _dopuni_ur0_racunom(ws, red, ur, br_rac, datum, vrsta, neto, pdv, rok):
    """Dopuni postojeći UR0 red podacima e-računa. Uplata (IZVADAK/PLAĆENO/datum plać./
    SREDSTVO) i formule (UKUPNO=K+J, DUGOVANJE=L-P) ostaju netaknute -> red postaje
    plaćeni račun (DUGUJE ≈ 0)."""
    ws.cell(red, 1).value = ur            # 0 -> pravi UR broj
    ws.cell(red, 2).value = br_rac        # broj računa
    ws.cell(red, 4).value = datum         # PRAVI datum računa (umjesto datuma uplate)
    if vrsta:
        ws.cell(red, 8).value = vrsta
    ws.cell(red, 10).value = neto         # TROŠAK (neto)
    ws.cell(red, 11).value = pdv          # PDV
    if rok:
        ws.cell(red, 13).value = rok


def _opis(r):
    return (f"{(r.get('supplierRegistrationName') or '')[:30]:30} | "
            f"{_broj(r.get('totalAmount')):9.2f} € | "
            f"{(r.get('documentDateInCet') or '')[:10]} | {r.get('invoiceNumber')}")


def izbornik():
    """Vrati 'prikazi' ili 'odradi'. Ako je SIMULACIJA uključena -> uvijek 'prikazi'."""
    if config.SIMULACIJA:
        return "prikazi"
    print("\nŠto želiš napraviti?")
    print("  1) SAMO PRIKAŽI što je novo   (ništa ne dira)")
    print("  2) ODRADI SVE                 (upis + razvrstavanje i povezivanje datoteka)")
    try:
        izbor = input("Izbor [1/2]: ").strip()
    except EOFError:
        return "prikazi"
    return "odradi" if izbor == "2" else "prikazi"


# ---------------------------------------------------------------------
#  Priprema (čitanje + razvrstavanje računa)
# ---------------------------------------------------------------------

def pripremi(ws, logger, stat, stanje):
    """Vrati (klase, downloads_sve, red_start, ur_start, zadnji_u_knjizi)."""
    zadnji_red, zadnji_ur, postojeci = excel_mod.analiziraj_postojece(ws, config.HEADER_RED)
    logger.info("Zadnji red: %s | zadnji UR: %s | već upisanih (po broju): %s",
                zadnji_red, zadnji_ur, len(postojeci))

    logger.info("Dohvaćam ULAZNE e-račune s Parre...")
    svi = parra_mod.dohvati_ulazne_racune(
        config.PARRA_API_TOKEN, config.PARRA_BUSINESS_ID, config.PARRA_BASE_URL,
        logger, config.PARRA_PAGE_SIZE, config.PARRA_MAX_STRANICA,
    )
    stat["skinuto"] = len(svi)

    # DATUM STARTA: makni račune prije OBRADJUJ_OD (to je ručni zaostatak, ne diramo)
    prije = [r for r in svi if (_parse_datum(r.get("documentDateInCet")) or config.OBRADJUJ_OD) < config.OBRADJUJ_OD]
    if prije:
        svi = [r for r in svi if r not in prije]
        logger.info("Preskačem %s računa prije starta (%s) — ručni zaostatak.",
                    len(prije), config.OBRADJUJ_OD)

    # VEĆ OBRAĐENI (zapamćeni): ne vraćaj ono što je već jednom upisano (i možda ručno obrisano)
    vec = [r for r in svi if r.get("id") in stanje["racuni"]]
    if vec:
        svi = [r for r in svi if r.get("id") not in stanje["racuni"]]
        logger.info("Preskačem %s već obrađenih računa (zapamćeno, neće se vraćati).", len(vec))

    preostali = parra_mod.filtriraj_nove(svi, postojeci)
    stat["vec_po_broju"] = len(svi) - len(preostali)

    logger.info("Učim nazive i vrste troška iz knjige...")
    wb_val = excel_mod.ucitaj_workbook_vrijednosti(config.EXCEL_PATH)
    history = vrsta_mod.nauci_iz_knjige(wb_val[config.GLAVNI_SHEET], config.HEADER_RED)
    imena = vrsta_mod.nauci_imena(wb_val[config.GLAVNI_SHEET], config.HEADER_RED)
    wb_val.close()

    # BROJ RAČUNA JE KLJUČ (dogovoreno s korisnikom): 'preostali' su računi čiji broj NIJE u knjizi
    # (filtriraj_nove) i koji su od datuma starta. Različit broj = SIGURNO NIJE duplikat — čak i
    # kod istog iznosa/datuma (mjesečni računi). Zato su SVI 'novo'. (Stari dedup po iznosu+datumu
    # je isključen jer je krivo označavao mjesečne račune kao 'mogući duplikat'.)
    klase = {"novo": [(r, None) for r in preostali], "neizvjesno": [], "duplikat": []}
    stat["duplikat"] = 0
    stat["neizvjesno"] = 0
    stat["novo"] = len(preostali)

    # Podsjetnik: zadnji račun s Parre koji je VEĆ u knjizi (najnoviji po datumu)
    novi_ids = {r.get("id") for r, _ in (klase["novo"] + klase["neizvjesno"])}
    u_knjizi = [r for r in svi if r.get("id") not in novi_ids]
    zadnji_u_knjizi = max(u_knjizi, key=lambda r: (r.get("documentDateInCet") or ""), default=None)

    # Skeniraj Downloads (sve datoteke: pdf + xml)
    downloads_sve = dat_mod.skeniraj_sve(config.FOLDER_DOWNLOADS)
    return klase, downloads_sve, zadnji_red + 1, zadnji_ur + 1, zadnji_u_knjizi, history, imena


# ---------------------------------------------------------------------
#  Razvrstavanje i povezivanje datoteka jednog računa
# ---------------------------------------------------------------------

def obradi_datoteke(ws, red, ur, br_rac, datum, downloads, logger, stat, mijenjaj, tiho=False,
                    ispravi_iznos=False):
    """Pronađi i razvrstaj datoteke za jedan račun.
    XML + Parrin PDF -> Parra/{mjesec}; originalni PDF -> pdf/ kao UR XXXX.pdf (+ link).
    'mijenjaj' = True znači stvarno premještaj/linkaj; False = samo ispiši što bi napravio.
    'tiho' = True ispisuje samo uspjehe (za prolaz kroz postojeće retke).
    'ispravi_iznos' = True: ako je nađeni original OTP leasing obrok-PDF, upiši
       'Sveukupno za platiti' (pun obrok) umjesto kamate s Parre. Samo za NOVE retke.
    """
    citaj = pdf_utils.procitaj_tekst_pdf if config.PDF_MATCH_SADRZAJ else None
    nadjeno = dat_mod.nadji_datoteke_racuna(br_rac, downloads, citaj)

    # 1) XML + Parrin PDF -> Parra/{mjesec}  (zapamti kamo je Parrin PDF otišao,
    #    da ga možemo iskoristiti kao zamjenu ako nema originalnog PDF-a)
    parra_pdf_final = None
    ima_parra_pdf = bool(nadjeno["parra_pdf"])
    parra_dat = nadjeno["xml"] + nadjeno["parra_pdf"]
    if parra_dat and datum:
        for d in parra_dat:
            if mijenjaj:
                try:
                    mapa = dat_mod.mjesec_mapa(config.FOLDER_PARRA, datum)
                    nova = dat_mod.spremi(d, mapa)
                    if d.get("ext") == ".pdf":
                        parra_pdf_final = nova
                    if d in downloads:
                        downloads.remove(d)
                    izvor = "(iz ZIP-a)" if d.get("zip") else ""
                    logger.info("   📁 %s → Parra/%s/ %s", d["ime"], dat_mod.HR_MJESECI.get(datum.month), izvor)
                    stat["parra_premjesteno"] += 1
                except Exception as e:
                    logger.error("   ⚠️ Ne mogu premjestiti %s: %s", d["ime"], e)
            elif not tiho:
                logger.info("   [prikaz] BIH premjestio %s → Parra/%s/", d["ime"],
                            dat_mod.HR_MJESECI.get(datum.month))

    # 2) Originalni PDF dobavljača -> pdf/ kao "UR XXXX.pdf" (+ link)
    novo_ime = f"UR {str(ur).zfill(4)}.pdf"
    cilj_pdf = os.path.join(config.FOLDER_PDF_FINAL, novo_ime)

    orig = nadjeno["original_pdf"]
    if orig:
        d = orig[0]
        if len(orig) > 1:
            logger.warning("   ⚠️ Nađeno %s mogućih originalnih PDF-ova za %s — uzimam '%s'.",
                           len(orig), br_rac, d["ime"])
        if mijenjaj:
            try:
                # OTP leasing: pun obrok ('Sveukupno za platiti') umjesto kamate s Parre.
                # Čitamo SIROVI tekst PDF-a (d["sadrzaj"] je očišćen — bez razmaka/zareza).
                if ispravi_iznos and d.get("putanja"):
                    obrok = dat_mod.sveukupno_za_platiti(pdf_utils.procitaj_tekst_pdf(d["putanja"]))
                    if obrok:
                        ws.cell(row=red, column=10).value = obrok  # TROŠAK = pun obrok
                        ws.cell(row=red, column=11).value = 0       # PDV (OTP leasing = 0)
                        logger.info("   💶 leasing: 'Sveukupno za platiti' %.2f€ (umjesto kamate)", obrok)
                dat_mod.spremi(d, config.FOLDER_PDF_FINAL, novo_ime)
                if d in downloads:
                    downloads.remove(d)
                excel_mod.postavi_hyperlink(ws, red, 1, cilj_pdf)
                logger.info("   ✅ originalni PDF '%s' → %s (+ link)", d["ime"], novo_ime)
                stat["pdf_spojen"] += 1
            except Exception as e:
                logger.error("   ⚠️ Ne mogu premjestiti/povezati %s: %s", d["ime"], e)
                stat["pdf_nije_nadjen"] += 1
        elif not tiho:
            logger.info("   [prikaz] BIH spremio '%s' kao %s i povezao.", d["ime"], novo_ime)
            stat["pdf_spojen"] += 1
    elif os.path.exists(cilj_pdf):
        # PDF već postoji u pdf/ pod tim UR brojem -> samo poveži
        if mijenjaj:
            excel_mod.postavi_hyperlink(ws, red, 1, cilj_pdf)
            logger.info("   ✅ %s već postoji — povezano.", novo_ime)
        elif not tiho:
            logger.info("   [prikaz] %s već postoji — BIH povezao.", novo_ime)
        stat["pdf_spojen"] += 1
    elif ima_parra_pdf:
        # NEMA originalnog PDF-a -> zamjena: KOPIRAJ Parrin PDF u pdf/ kao "UR XXXX.pdf"
        if mijenjaj and parra_pdf_final:
            try:
                shutil.copy2(parra_pdf_final, cilj_pdf)
                excel_mod.postavi_hyperlink(ws, red, 1, cilj_pdf)
                logger.info("   📄 nema originala — Parrin PDF kopiran kao %s (+ link)", novo_ime)
                stat["pdf_spojen"] += 1
            except Exception as e:
                logger.error("   ⚠️ Ne mogu kopirati Parrin PDF kao %s: %s", novo_ime, e)
                stat["pdf_nije_nadjen"] += 1
        elif not mijenjaj and not tiho:
            logger.info("   [prikaz] nema originala — BIH kopirao Parrin PDF kao %s.", novo_ime)
            stat["pdf_spojen"] += 1
    elif not tiho:
        logger.info("   ❌ ni original ni Parrin PDF nije nađen (skini ga s Parre u Downloads).")
        stat["pdf_nije_nadjen"] += 1


# ---------------------------------------------------------------------
#  Upis novih + povezivanje postojećih
# ---------------------------------------------------------------------

def _vrsta_za_eracun(r, history, logger):
    """Predloži VRSTU TROŠKA: prvo iz povijesti; ako ne zna, dohvati stavke/KPD računa."""
    partner = r.get("supplierRegistrationName")
    v = vrsta_mod.predlozi_vrstu(partner, history)   # povijest (jeftino, bez API-ja)
    if v:
        return v
    detalj = parra_mod.dohvati_detalj(config.PARRA_API_TOKEN, config.PARRA_BUSINESS_ID,
                                      config.PARRA_BASE_URL, r.get("id"), logger)
    if not detalj:
        return None
    linije = detalj.get("invoiceLines") or []
    tekst = " ".join(f"{l.get('name', '')} {l.get('description') or ''}" for l in linije)
    kpd = (linije[0].get("kpdMark") if linije else "") or ""
    return vrsta_mod.predlozi_vrstu(partner, history, tekst, kpd)


def upisi_nove(ws, klase, downloads, red_start, ur_start, logger, stat, pisi, history=None,
               imena=None, stanje=None, ceka_racun=None):
    """Upiši OČITO NOVE i NEIZVJESNE račune.
    Vrati (promijenjeno, [], upisani_racuni) — upisani za spajanje s izvodima u istom pokretanju.
    Ako za račun već postoji UR0 uplata (deferno) — dopuni je umjesto novog reda.
    """
    red, ur = red_start, ur_start
    promijenjeno = False
    upisani = []   # index-zapisi za spajanje (row, ur, racun, dob, iznos, datum...)

    # Upisujemo NOVE i NEIZVJESNE (isti iznos+drugi datum = mjesečni račun, NIJE duplikat
    # jer se broj računa ne poklapa). Sortirano kronološki da UR ide redom.
    za_upis = [r for r, _ in klase["novo"]] + [r for r, _ in klase["neizvjesno"]]
    za_upis.sort(key=lambda r: (r.get("documentDateInCet") or ""))

    if za_upis:
        logger.info("--- NOVI računi za upis (%s, uklj. %s 'mogući duplikat') ---",
                    len(za_upis), len(klase["neizvjesno"]))
    for r in za_upis:
        try:
            br_rac = str(r.get("invoiceNumber", "")).strip()
            datum = _parse_datum(r.get("documentDateInCet"))
            dobavljac = vrsta_mod.kanonsko_ime(r.get("supplierRegistrationName"), imena or {})
            total = _broj(r.get("totalAmount"))
            neto = _broj(r.get("totalAmountWithoutTax"))
            pdv = _broj(r.get("totalTaxAmount"))
            # UREDBA (umanjenje cijene struje): ako je 'za platiti' (payableAmount) manji od
            # ukupnog -> preračunaj osnovicu/PDV iz umanjenog iznosa po istoj stopi (kao knjigovođa).
            _dob_blob = ((dobavljac or "") + " " + (r.get("supplierRegistrationName") or "")).lower()
            if total and neto and any(k in _dob_blob for k in ("energij", "elektra", "plin")):
                _det = parra_mod.dohvati_detalj(config.PARRA_API_TOKEN, config.PARRA_BUSINESS_ID,
                                                config.PARRA_BASE_URL, r.get("id"), logger)
                _pay = (_det or {}).get("totals", {}).get("payableAmount")
                if _pay is not None and _pay < total - 0.01:
                    neto = round(_pay * neto / total, 2)
                    pdv = round(_pay - neto, 2)
                    total = round(_pay, 2)
                    logger.info("   ⚡ umanjeno po Uredbi: za platiti %.2f € (osnovica %.2f + PDV %.2f)",
                                total, neto, pdv)
            logger.info("-> UR %04d | %s", ur, _opis(r))

            vrsta = _vrsta_za_eracun(r, history, logger) if (pisi and history is not None) else None

            # DEFERNO: postoji li UR0 uplata (od starta) za ovaj račun? (jednoznačno)
            cilj = (_deferno_kandidat(ocisti(dobavljac), total, datum, ceka_racun)
                    if (pisi and ceka_racun) else None)
            if cilj:
                if pisi:
                    _dopuni_ur0_racunom(ws, cilj["row"], ur, br_rac, datum, vrsta,
                                        neto, pdv, _parse_datum(r.get("dueDate")))
                    cilj["iskoristen"] = True
                    promijenjeno = True
                    stat["upisano"] += 1
                    stat["deferno"] += 1
                    if stanje is not None and r.get("id") is not None:
                        stanje["racuni"].add(r.get("id"))
                    obradi_datoteke(ws, cilj["row"], ur, br_rac, datum, downloads, logger, stat,
                                    mijenjaj=config.STVARNO_DATOTEKE, ispravi_iznos=True)
                logger.info("   🔗 deferno: UR%04d (%s, %.2f€) dopunjen na postojeću uplatu (red %s)",
                            ur, (dobavljac or "")[:20], total, cilj["row"])
                ur += 1
                continue   # NE dodajemo novi red — red ostaje za sljedeći račun

            # NORMALNO: novi red na dnu
            if pisi:
                podaci = {
                    1: ur, 2: br_rac, 4: datum, 5: dobavljac,
                    8: vrsta,
                    10: neto, 11: pdv,
                    12: total, 13: _parse_datum(r.get("dueDate")),
                }
                excel_mod.upisi_vrijednosti(ws, red, podaci, stil_izvor_red=red - 1)
                promijenjeno = True
                stat["upisano"] += 1
                if stanje is not None and r.get("id") is not None:
                    stanje["racuni"].add(r.get("id"))   # zapamti da je obrađen
                # zapamti za spajanje s izvodima (čeka plaćanje)
                upisani.append({"row": red, "ur": ur, "racun": br_rac, "dob": dobavljac,
                                "dob_clean": ocisti(dobavljac), "iznos": total,
                                "datum": datum, "ima_izvadak": False, "vrsta": vrsta})

            # Razvrstaj/poveži datoteke (premještaj samo u pravom radu)
            obradi_datoteke(ws, red, ur, br_rac, datum, downloads, logger, stat,
                            mijenjaj=config.STVARNO_DATOTEKE, ispravi_iznos=True)
            red += 1
            ur += 1
        except Exception as e:
            stat["greske"] += 1
            logger.error("   ⚠️ Greška kod %s: %s — preskačem.", r.get("invoiceNumber"), e)
            red += 1
            ur += 1

    return promijenjeno, [], upisani


def povezi_postojece(ws, downloads, logger, stat, pisi):
    """Za postojeće retke koji imaju UR i RAČUN ali NEMAJU link na PDF — pokušaj povezati."""
    if not pisi:
        return False
    promijenjeno = False
    for i in range(config.HEADER_RED + 1, ws.max_row + 1):
        celija_ur = ws.cell(row=i, column=1)
        ur_val = celija_ur.value
        br_rac = ws.cell(row=i, column=2).value
        # samo retci s pravim UR brojem, brojem računa i BEZ postojećeg linka
        if not br_rac or celija_ur.hyperlink is not None:
            continue
        try:
            ur = int(str(ur_val).split(".")[0])
        except (ValueError, TypeError):
            continue
        if ur <= 0:
            continue
        datum = _datum_celije(ws.cell(row=i, column=4).value)
        # ne diramo stari ručni dio (prije datuma starta)
        if datum and datum < config.OBRADJUJ_OD:
            continue
        prije = stat["pdf_spojen"]
        obradi_datoteke(ws, i, ur, str(br_rac), datum, downloads, logger, stat,
                        mijenjaj=config.STVARNO_DATOTEKE, tiho=True)
        if stat["pdf_spojen"] > prije:
            promijenjeno = True
    return promijenjeno


def _obogati_obroke(indeks, logger):
    """Za leasing račune koji ČEKAJU (bez izvatka): pročitaj BROJ OBROKA iz XML-a u Parra
    mapi i upiši ga u indeks (za spajanje po obroku). Tiho — ako ne nađe, ostaje None."""
    import glob
    leasing = [r for r in indeks["racuni"]
               if not r.get("ima_izvadak") and "leasing" in (r.get("dob") or "").lower()]
    if not leasing:
        return
    try:
        xmls = glob.glob(os.path.join(config.FOLDER_PARRA, "**", "*.xml"), recursive=True)
    except Exception:
        return
    meta = [(p, os.path.basename(p)) for p in xmls]
    for r in leasing:
        br = str(r.get("racun") or "")
        for p, ime in meta:
            if dat_mod.ime_odgovara_racunu(br, {"ime": ime, "ime_clean": dat_mod.ocisti(ime)}):
                try:
                    r["obrok"] = spajanje_mod.broj_obroka(
                        open(p, encoding="utf-8", errors="ignore").read())
                except Exception:
                    r["obrok"] = None
                break
    n = sum(1 for r in leasing if r.get("obrok") is not None)
    logger.info("   leasing: broj obroka iz XML-a za %s/%s računa koji čekaju", n, len(leasing))


def obradi_izvode_tok(ws, logger, stat, pisi, history, imena, extra_racuni=None, stanje=None):
    """Obradi bankovne izvode (HPB + RBA): spoji s računima + UR0 redovi. Vrati promijenjeno."""
    logger.info("-" * 64)
    logger.info("IZVODI (HPB + RBA)")
    # Indeks + popis već unesenih iz data_only kopije (zbog formula UKUPNO/PDV)
    wb_val = excel_mod.ucitaj_workbook_vrijednosti(config.EXCEL_PATH)
    wsv = wb_val[config.GLAVNI_SHEET]
    indeks = spajanje_mod.izgradi_indeks(wsv, config.HEADER_RED)
    uneseni = izvod_mod.uneseni_izvodi(wsv, config.HEADER_RED)
    wb_val.close()
    # dodaj račune upisane u OVOM pokretanju (čekaju plaćanje) da ih izvodi mogu spojiti
    if extra_racuni:
        indeks["racuni"].extend(extra_racuni)
    _obogati_obroke(indeks, logger)   # leasing: dodaj broj obroka iz XML-a (za spajanje po obroku)

    zadnji_red, _, _ = excel_mod.analiziraj_postojece(ws, config.HEADER_RED)
    folderi = [("HPB", config.FOLDER_IZVODI_HPB), ("RBA", config.FOLDER_IZVODI_RBA)]
    promijenjeno, _ = izvod_mod.obradi_izvode(
        ws, indeks, uneseni, folderi, logger, stat, pisi, config.STVARNO_DATOTEKE,
        zadnji_red, history=history, obradjuj_od=config.IZVODI_OD, imena=imena, stanje=stanje)
    return promijenjeno


# ---------------------------------------------------------------------
#  FAZA 2 — putni nalozi (čita iz mjesečnog PN Excela)
# ---------------------------------------------------------------------

def obradi_putne(ws, pn_excel_path, logger, stat, pisi, imena=None, stanje=None):
    """Pročitaj putne naloge iz mjesečnog PN Excela i upiši ih kao retke (VRSTA='dnevnice').
    DJELATNIK u kolonu F, RAČUN='pn XXX', TROŠAK=iznos za isplatu, PLAĆENO ostaje prazno
    (popunit će se kad dnevnice budu plaćene na izvodu). Vrati (promijenjeno, koliko_upisano)."""
    try:
        nalozi = putni_mod.procitaj_naloge(pn_excel_path)
    except Exception as e:
        logger.error("Ne mogu pročitati PN Excel %s: %s", pn_excel_path, e)
        return False, 0
    # upisuj po rastućem PN broju -> i UR brojevi idu redom (od manjeg prema većem)
    nalozi.sort(key=lambda n: int("".join(ch for ch in n["broj"] if ch.isdigit()) or 0))
    logger.info("-" * 64)
    logger.info("PUTNI NALOZI — %s naloga u: %s", len(nalozi), os.path.basename(pn_excel_path))

    # već upisani 'pn XXX' (iz knjige + iz pamćenja, da se obrisani ne vraćaju)
    postojeci = set()
    for i in range(config.HEADER_RED + 1, ws.max_row + 1):
        rac = str(ws.cell(row=i, column=2).value or "").lower().replace(" ", "")
        if rac.startswith("pn"):
            postojeci.add(rac)
    if stanje is not None:
        postojeci |= {k.replace(" ", "") for k in stanje.get("putni", set())}

    zadnji_red, zadnji_ur, _ = excel_mod.analiziraj_postojece(ws, config.HEADER_RED)
    red, ur = zadnji_red + 1, zadnji_ur + 1
    promijenjeno = False
    upisano = 0
    for n in nalozi:
        rac = f"pn {n['broj']}"
        if rac.replace(" ", "") in postojeci:
            continue  # već upisan
        djel = (n["djelatnik"] or "").lower()   # knjiga koristi mala slova (npr. 'ivan horvat')
        logger.info("-> UR %04d | %s | %s | %.2f€", ur, rac, djel[:20], n["iznos"])
        if pisi:
            podaci = {1: ur, 2: rac, 4: n["datum"], 6: djel,
                      8: "dnevnice", 10: n["iznos"], 11: 0}
            excel_mod.upisi_vrijednosti(ws, red, podaci, stil_izvor_red=red - 1)
            promijenjeno = True
            if stanje is not None:
                stanje["putni"].add(rac)
        red += 1
        ur += 1
        upisano += 1
    logger.info("Putnih naloga upisano: %s", upisano)
    return promijenjeno, upisano


# ---------------------------------------------------------------------
#  Glavni tok
# ---------------------------------------------------------------------

def _novi_stat():
    return {"skinuto": 0, "vec_po_broju": 0, "duplikat": 0, "neizvjesno": 0, "novo": 0,
            "upisano": 0, "pdf_spojen": 0, "pdf_nije_nadjen": 0,
            "parra_premjesteno": 0, "greske": 0,
            # izvodi:
            "izvoda_obradeno": 0, "spojeno": 0, "dnevnice": 0, "ur0_redova": 0,
            "preskoceno_uplata": 0, "deferno": 0, "linkovi_dodani": 0,
            # za prijateljski status (GUI):
            "zadnji_racun": None}


def _ur_iz_imena(f):
    """Iz naziva PDF-a izvuci UR broj. Hvata 'UR 0492.pdf' i 'UR 0492_PN 060.pdf' -> 492.
    None ako ne počinje s 'UR' + broj ili nije .pdf."""
    if not f.lower().endswith(".pdf"):
        return None
    ostatak = f.strip()
    if not ostatak[:2].lower() == "ur":
        return None
    ostatak = ostatak[2:].lstrip()          # nakon 'UR' i razmaka
    broj = ""
    for ch in ostatak:
        if ch.isdigit():
            broj += ch
        else:
            break                            # broj završava na prvom ne-znamenki
    return int(broj) if broj else None


def dopuni_hyperlinkove(ws, logger, stat):
    """Za UR retke kojima FALI klik-poveznica: ako u pdf mapi postoji datoteka koja počinje
    s tim UR brojem (npr. 'UR 0492.pdf' ili 'UR 0492_PN 060.pdf'), dodaj plavu poveznicu.
    Vrati True ako je nešto dodano."""
    # jednom pročitaj pdf mapu i složi mapu: UR broj -> naziv datoteke
    mapa = {}
    try:
        for f in sorted(os.listdir(config.FOLDER_PDF_FINAL)):
            ur = _ur_iz_imena(f)
            if ur is not None:
                mapa.setdefault(ur, f)       # prvi po abecedi ako ih je više
    except OSError as e:
        logger.error("Ne mogu pročitati pdf mapu: %s", e)
        return False

    dodano = 0
    ur_u_knjizi = set()
    bez_pdfa = []
    for i in range(config.HEADER_RED + 1, ws.max_row + 1):
        c = ws.cell(row=i, column=1)
        ur = c.value
        if isinstance(ur, float) and ur.is_integer():
            ur = int(ur)
        if not isinstance(ur, int) or ur <= 0:
            continue
        ur_u_knjizi.add(ur)
        if c.hyperlink is not None:
            continue                         # već ima poveznicu
        f = mapa.get(ur)
        if f:
            excel_mod.postavi_hyperlink(ws, i, 1, os.path.join(config.FOLDER_PDF_FINAL, f))
            dodano += 1
            logger.info("   🔗 UR %04d → %s", ur, f)
        else:
            bez_pdfa.append(ur)              # nema ni datoteke ni poveznice
    stat["linkovi_dodani"] = dodano
    if dodano:
        logger.info("Dopunjeno poveznica koje su falile: %s", dodano)

    # UPOZORENJE na mogući tipfeler: PDF u mapi čiji UR broj NE postoji u knjizi
    visak = sorted(u for u in mapa if u not in ur_u_knjizi)
    if visak:
        logger.warning("⚠️ PDF-ovi u pdf mapi BEZ retka u knjizi (tipfeler u nazivu?): %s",
                       ", ".join(f"UR {u:04d} ({mapa[u]})" for u in visak[:10]))

    # INFO: redci kojima još fali PDF/poveznica (što još treba skenirati)
    if bez_pdfa:
        prikaz = ", ".join(f"{u:04d}" for u in bez_pdfa[:25])
        logger.info("📄 Redaka bez PDF poveznice: %s%s → UR %s",
                    len(bez_pdfa), " (prvih 25)" if len(bez_pdfa) > 25 else "", prikaz)
    return dodano > 0


def pokreni_obradu(nacin, logger):
    """JEZGRA obrade (bez konzolnog izbornika) — zove je i konzola i GUI.
    'nacin' = 'prikazi' (ništa ne dira) ili 'odradi' (upiše + premjesti). Vrati stat."""
    logger.info("=" * 64)
    logger.info("URA URA — ulazni e-računi s Parre")
    logger.info("Vrijeme: %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    if config.KORISTI_TEST_EXCEL:
        logger.info("⚠️  TEST NAČIN: radim na KOPIJI; datoteke se NE premještaju.")
    logger.info("Excel: %s", config.EXCEL_PATH)

    # Smijemo li pisati u Excel / premještati datoteke?
    config.PISI_EXCEL = (nacin == "odradi") and not config.SIMULACIJA
    config.STVARNO_DATOTEKE = config.PISI_EXCEL and (
        not config.KORISTI_TEST_EXCEL or getattr(config, "PRISILI_DATOTEKE", False))
    logger.info("Način rada: %s", "ODRADI SVE" if config.PISI_EXCEL else "SAMO PRIKAŽI")
    logger.info("=" * 64)

    stat = _novi_stat()

    if not config.PARRA_API_TOKEN:
        logger.error("Nema PARRA_API_TOKEN u .env datoteci. Prekidam.")
        return stat
    if not os.path.exists(config.EXCEL_PATH):
        # Razlikuj "server nije spojen" od "datoteka fali" — da poruka odmah kaže što napraviti
        disk = os.path.splitdrive(config.EXCEL_PATH)[0]      # npr. 'Z:'
        if disk and not os.path.exists(disk + "\\"):
            logger.error("SERVER NIJE DOSTUPAN: disk %s nije spojen. Provjeri mrežu/VPN "
                         "ili otvori %s u Exploreru pa pokušaj ponovno.", disk, disk)
        else:
            logger.error("Excel ne postoji: %s", config.EXCEL_PATH)
        return stat

    logger.info("Učitavam Excel...")
    try:
        wb = excel_mod.ucitaj_workbook(config.EXCEL_PATH)
        ws = wb[config.GLAVNI_SHEET]
    except Exception as e:
        logger.error("Ne mogu učitati Excel: %s", e)
        return stat

    # Predložak stila po stupcu (iz postojećih redova) — da novi redovi budu IDENTIČNI
    excel_mod.pripremi_stil(ws, config.HEADER_RED)

    stanje = stanje_mod.ucitaj(config.STANJE_PATH)   # što je već obrađeno (da se ne vraća)
    klase, downloads, red_start, ur_start, zadnji, history, imena = pripremi(ws, logger, stat, stanje)

    # Podsjetnik
    if zadnji:
        stat["zadnji_racun"] = (f"{zadnji.get('invoiceNumber')}, "
                                f"{zadnji.get('supplierRegistrationName')}, "
                                f"{(zadnji.get('documentDateInCet') or '')[:10]}")
        logger.info("📌 Zadnji račun s Parre koji je u knjizi: %s", stat["zadnji_racun"])
    logger.info("Novih: %s | neizvjesnih(mogući duplikat): %s | duplikata: %s | već po broju: %s",
                stat["novo"], stat["neizvjesno"], stat["duplikat"], stat["vec_po_broju"])

    if not (klase["novo"] or klase["neizvjesno"]):
        logger.info("Nema novih računa.")

    # BACKUP prije ikakvog upisa
    if config.PISI_EXCEL:
        try:
            backup_mod.napravi_backup(config.EXCEL_PATH, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
        except Exception as e:
            logger.error("Backup NIJE uspio (%s). Iz sigurnosti prekidam.", e)
            return stat

    # 1) E-RAČUNI  (uz deferno spajanje: dopuni UR0 uplatu koja je čekala račun)
    ceka_racun = _indeks_ceka_racun(ws, config.OBRADJUJ_OD)
    promj1, _, upisani_racuni = upisi_nove(ws, klase, downloads, red_start, ur_start, logger, stat,
                                           config.PISI_EXCEL, history, imena, stanje, ceka_racun)
    promj2 = povezi_postojece(ws, downloads, logger, stat, config.PISI_EXCEL)

    # 2) IZVODI (HPB + RBA) — uklj. račune upisane u OVOM pokretanju (za spajanje)
    try:
        promj3 = obradi_izvode_tok(ws, logger, stat, config.PISI_EXCEL, history, imena,
                                   upisani_racuni, stanje)
    except Exception as e:
        promj3 = False
        logger.error("Greška kod obrade izvoda: %s", e)

    # 3) Dopuni klik-poveznice koje fale (npr. ručno spremljeni skenovi 'UR XXXX.pdf')
    promj4 = False
    if config.PISI_EXCEL:
        try:
            promj4 = dopuni_hyperlinkove(ws, logger, stat)
        except Exception as e:
            logger.error("Greška kod dopune poveznica: %s", e)

    # Spremanje
    if config.PISI_EXCEL and (promj1 or promj2 or promj3 or promj4):
        try:
            excel_mod.spremi(wb, config.EXCEL_PATH, logger)
            stanje_mod.spremi(stanje, config.STANJE_PATH)   # zapamti obrađeno (tek nakon uspjeha)
        except Exception as e:
            logger.error("Spremanje nije uspjelo: %s (backup je napravljen)", e)
            return stat
    elif not config.PISI_EXCEL:
        logger.info("(Prikaz — Excel nije spremljen.)")

    _izvjestaj(logger, stat)
    return stat


def pokreni_putne(godina, mjesec, logger):
    """Obradi putne naloge za zadani mjesec (iz PN Excela). Vrati koliko upisano.
    Zasebna naredba (gumb) — pokreće se jednom mjesečno."""
    pn_path = config.pn_excel_putanja(godina, mjesec)
    logger.info("=" * 64)
    logger.info("PUTNI NALOZI — %02d-%s", mjesec, godina)
    logger.info("Excel: %s", pn_path)
    if not os.path.exists(pn_path):
        logger.error("Ne nalazim PN Excel za taj mjesec: %s", pn_path)
        return 0
    if not os.path.exists(config.EXCEL_PATH):
        logger.error("Knjiga ne postoji: %s", config.EXCEL_PATH)
        return 0

    pisi = not config.SIMULACIJA
    wb = excel_mod.ucitaj_workbook(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    excel_mod.pripremi_stil(ws, config.HEADER_RED)
    stanje = stanje_mod.ucitaj(config.STANJE_PATH)

    if pisi:
        try:
            backup_mod.napravi_backup(config.EXCEL_PATH, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
        except Exception as e:
            logger.error("Backup NIJE uspio (%s). Prekidam.", e)
            return 0

    promijenjeno, upisano = obradi_putne(ws, pn_path, logger, {}, pisi, imena=None, stanje=stanje)
    if pisi and promijenjeno:
        try:
            excel_mod.spremi(wb, config.EXCEL_PATH, logger)
            stanje_mod.spremi(stanje, config.STANJE_PATH)
        except Exception as e:
            logger.error("Spremanje nije uspjelo: %s (backup je napravljen)", e)
            return 0
    elif not pisi:
        logger.info("(SIMULACIJA — ništa nije spremljeno.)")
    logger.info("=" * 64)
    return upisano


def izvjesce_racuni_bez_izvoda(od_datuma=None):
    """Računi kojima fali izvod (nisu plaćeni). Vrati listu (samo čita knjigu)."""
    wb = excel_mod.ucitaj_workbook_vrijednosti(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    rez = izvjestaji_mod.racuni_bez_izvoda(ws, config.HEADER_RED, od_datuma)
    wb.close()
    return rez


def izvjesce_izvodi_bez_racuna(od_datuma=None):
    """Uplate s izvoda (UR0) kojima fale računi, bez trajnih. Vrati listu (samo čita)."""
    wb = excel_mod.ucitaj_workbook_vrijednosti(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    rez = izvjestaji_mod.izvodi_bez_racuna(ws, config.HEADER_RED, od_datuma)
    wb.close()
    return rez


def izvjesce_treba_potvrdu(od_datuma=None):
    """Računi kojima fali VRSTA TROŠKA (treba potvrdu). Vrati listu (samo čita knjigu)."""
    wb = excel_mod.ucitaj_workbook_vrijednosti(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    rez = izvjestaji_mod.racuni_bez_vrste(ws, config.HEADER_RED, od_datuma)
    wb.close()
    return rez


def skeniraj(logger):
    """PREGLED: dohvati s Parre + klasificiraj, BEZ upisa. Vrati 'sesiju' (drži otvorenu
    knjigu i kandidate) za ekran potvrde. None ako ne može."""
    logger.info("=" * 64)
    logger.info("PREGLED — dohvaćam s Parre i razvrstavam…")
    config.PISI_EXCEL = False
    config.STVARNO_DATOTEKE = False
    stat = _novi_stat()
    if not config.PARRA_API_TOKEN:
        logger.error("Nema PARRA_API_TOKEN u .env. Prekidam.")
        return None
    if not os.path.exists(config.EXCEL_PATH):
        logger.error("Knjiga ne postoji: %s", config.EXCEL_PATH)
        return None
    wb = excel_mod.ucitaj_workbook(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    excel_mod.pripremi_stil(ws, config.HEADER_RED)
    stanje = stanje_mod.ucitaj(config.STANJE_PATH)
    klase, downloads, red_start, ur_start, zadnji, history, imena = pripremi(ws, logger, stat, stanje)

    if zadnji:
        stat["zadnji_racun"] = (f"{zadnji.get('invoiceNumber')}, "
                                f"{zadnji.get('supplierRegistrationName')}, "
                                f"{(zadnji.get('documentDateInCet') or '')[:10]}")

    za = ([(r, "novo") for r, _ in klase["novo"]] +
          [(r, "mogući duplikat") for r, _ in klase["neizvjesno"]])
    za.sort(key=lambda t: (t[0].get("documentDateInCet") or ""))
    kandidati = []
    for idx, (r, tip) in enumerate(za):
        kandidati.append({
            "idx": idx, "tip": tip,
            "broj": str(r.get("invoiceNumber") or ""),
            "dob": r.get("supplierRegistrationName") or "",
            "iznos": _broj(r.get("totalAmount")),
            "datum": (r.get("documentDateInCet") or "")[:10],
            "record": r,
        })
    logger.info("Za pregled: %s (novih %s + mogući duplikat %s)",
                len(kandidati), stat["novo"], stat["neizvjesno"])
    return {"wb": wb, "ws": ws, "downloads": downloads, "red_start": red_start,
            "ur_start": ur_start, "history": history, "imena": imena,
            "stanje": stanje, "stat": stat, "kandidati": kandidati}


def upisi_odabrane(sesija, odabrani_idx, logger):
    """Upiši SAMO odabrane kandidate (skup idx-eva), pa obradi izvode, pa spremi.
    Koristi 'sesiju' iz skeniraj(). Vrati stat."""
    config.PISI_EXCEL = not config.SIMULACIJA
    config.STVARNO_DATOTEKE = config.PISI_EXCEL and (
        not config.KORISTI_TEST_EXCEL or getattr(config, "PRISILI_DATOTEKE", False))
    ws, wb = sesija["ws"], sesija["wb"]
    stat, stanje = sesija["stat"], sesija["stanje"]
    odabrani = [k["record"] for k in sesija["kandidati"] if k["idx"] in set(odabrani_idx)]
    logger.info("=" * 64)
    logger.info("Upisujem %s odabranih računa (od %s ponuđenih)…",
                len(odabrani), len(sesija["kandidati"]))

    if config.PISI_EXCEL:
        try:
            backup_mod.napravi_backup(config.EXCEL_PATH, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
        except Exception as e:
            logger.error("Backup NIJE uspio (%s). Prekidam.", e)
            return stat

    klase_sel = {"novo": [(r, None) for r in odabrani], "neizvjesno": []}
    ceka_racun = _indeks_ceka_racun(ws, config.OBRADJUJ_OD)
    promj1, _, upisani = upisi_nove(ws, klase_sel, sesija["downloads"], sesija["red_start"],
                                    sesija["ur_start"], logger, stat, config.PISI_EXCEL,
                                    sesija["history"], sesija["imena"], stanje, ceka_racun)
    promj2 = povezi_postojece(ws, sesija["downloads"], logger, stat, config.PISI_EXCEL)
    try:
        promj3 = obradi_izvode_tok(ws, logger, stat, config.PISI_EXCEL,
                                   sesija["history"], sesija["imena"], upisani, stanje)
    except Exception as e:
        promj3 = False
        logger.error("Greška kod obrade izvoda: %s", e)

    if config.PISI_EXCEL and (promj1 or promj2 or promj3):
        try:
            excel_mod.spremi(wb, config.EXCEL_PATH, logger)
            stanje_mod.spremi(stanje, config.STANJE_PATH)
        except Exception as e:
            logger.error("Spremanje nije uspjelo: %s (backup je napravljen)", e)
            return stat
    _izvjestaj(logger, stat)
    return stat


def _nadji_uplatu(bruto, datum, uplate):
    """Pronađi UR0 uplatu s izvoda (bez računa) po iznosu (±tolerancija) i datumu (±14 dana).
    'bruto' je broj, 'datum' je date (ili None — tada se datum ne provjerava)."""
    if not bruto:
        return None
    najbolji, najd = None, 999
    for u in uplate:
        if abs((u["placeno"] or 0) - bruto) > config.DEDUP_TOL_IZNOS:
            continue
        ud = u["datum"].date() if hasattr(u["datum"], "date") else None
        dd = abs((ud - datum).days) if (datum and ud) else 0
        if dd <= 14 and dd < najd:
            najd, najbolji = dd, u
    return najbolji


def _spoji_fotku_s_uplatom(o, uplate):
    """Pronađi uplatu koja odgovara fotki po OCR-om pročitanom bruto-u i datumu."""
    return _nadji_uplatu(o.get("bruto"), _parse_datum_hr(o.get("datum")), uplate)


def skeniraj_fotke(logger):
    """Pripremi fotke za obradu — BEZ upfront OCR-a (lijeno: OCR svake tek kad dođe na ekran,
    pa se prozor otvori odmah). Vrati 'sesiju' (drži otvorenu knjigu). None ako nema fotki."""
    folder = config.FOLDER_FOTKE
    slike = fotke_mod.popis_fotki(folder)
    logger.info("=" * 64)
    logger.info("FOTKANI RAČUNI — %s u: %s (OCR po potrebi)", len(slike), folder)
    if not slike:
        logger.info("Nema fotki u mapi.")
        return None
    wbv = excel_mod.ucitaj_workbook_vrijednosti(config.EXCEL_PATH)
    uplate = izvjestaji_mod.izvodi_bez_racuna(wbv[config.GLAVNI_SHEET], config.HEADER_RED, None)
    wbv.close()
    wb = excel_mod.ucitaj_workbook(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    excel_mod.pripremi_stil(ws, config.HEADER_RED)
    stanje = stanje_mod.ucitaj(config.STANJE_PATH)
    kandidati = [{"slika": s, "ocr": None, "match": None} for s in slike]   # OCR kasnije, lijeno
    return {"wb": wb, "ws": ws, "stanje": stanje, "kandidati": kandidati,
            "uplate": uplate, "backup_done": False}


def ocr_fotke_jedan(sesija, idx, logger):
    """OCR + pokušaj spajanja na uplatu za JEDNU fotku (lijeno, kad dođe na red). Sprema u kandidat."""
    k = sesija["kandidati"][idx]
    if k["ocr"] is not None:
        return k
    try:
        o = ocr_mod.procitaj_racun(k["slika"])
    except Exception as e:
        logger.error("OCR greška %s: %s", os.path.basename(k["slika"]), e)
        o = {"datoteka": os.path.basename(k["slika"]), "broj": None, "datum": None,
             "bruto": None, "osnovica": None, "pdv": None, "tekst": ""}
    k["ocr"] = o
    k["match"] = _spoji_fotku_s_uplatom(o, sesija.get("uplate") or [])
    logger.info("  %s %s | bruto=%s | %s", o.get("izvor", "ocr").upper(),
                os.path.basename(k["slika"]), o.get("bruto"),
                (f"spojen na izvod (red {k['match']['red']})" if k["match"] else "bez para"))
    return k


def upisi_fotku(sesija, idx, polja, logger):
    """Upiši jednu fotku (potvrđena polja s ekrana): dopuni uplatu s izvoda ili novi red,
    pretvori sliku u 'UR XXXX.pdf' i poveži. 'polja' = {broj, datum, dobavljac, bruto,
    osnovica, pdv, vozilo, vrsta}. Vrati dodijeljeni UR broj."""
    ws, wb, stanje = sesija["ws"], sesija["wb"], sesija["stanje"]

    # DUPLIKAT PO BROJU + DATUMU: isti broj računa NIJE dovoljan za preskakanje jer neki
    # troškovi (npr. redovne bankovne kamate) uvijek imaju isti broj, a različit datum. Zato
    # preskačemo SAMO ako se poklapaju broj I datum. Isti broj + novi datum = upiši.
    def _kljuc_broja(s):
        return "".join(ch for ch in str(s or "").lower() if ch.isalnum())
    nb = _kljuc_broja(polja.get("broj"))
    nd = polja.get("datum")   # date objekt (ili None ako datum nije pročitan)
    if nb and not getattr(config, "DEMO", False):   # u DEMO načinu ne pamtimo (za demo/snimanje)
        for i in range(config.HEADER_RED + 1, ws.max_row + 1):
            rb = ws.cell(row=i, column=2).value
            if rb and _kljuc_broja(rb) == nb:
                c4 = ws.cell(row=i, column=4).value
                rd = _datum_celije(c4) or (_parse_datum_hr(c4) if isinstance(c4, str) else None)
                if nd is None or rd == nd:
                    post_ur = ws.cell(row=i, column=1).value
                    logger.info("⏭️ Fotka — broj %s + datum %s već postoji (UR %s) — NE upisujem.",
                                polja.get("broj"), nd, post_ur)
                    return {"duplikat": True, "ur": post_ur, "broj": polja.get("broj")}

    pisi = not config.SIMULACIJA
    premjesti = pisi and (not config.KORISTI_TEST_EXCEL or getattr(config, "PRISILI_DATOTEKE", False))
    if pisi and not sesija["backup_done"]:
        backup_mod.napravi_backup(config.EXCEL_PATH, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
        sesija["backup_done"] = True

    k = sesija["kandidati"][idx]

    # SPAJANJE S UPLATOM po POTVRĐENIM podacima (ne samo OCR-u!): korisnica često ručno
    # upiše/ispravi bruto i datum kad ih OCR ne pročita — zato ovdje PONOVNO tražimo
    # odgovarajuću uplatu s tim (točnim) vrijednostima. Inače bi ručno popunjene fotke
    # uvijek završile kao novi red BEZ veze s izvodom.
    if polja.get("bruto") is not None:
        k["match"] = _nadji_uplatu(polja["bruto"], polja.get("datum"), sesija.get("uplate") or [])
        if k["match"]:
            logger.info("   🔗 potvrđeni bruto %.2f → spojen na uplatu s izvoda (red %s)",
                        polja["bruto"], k["match"]["red"])

    _, zadnji_ur, _ = excel_mod.analiziraj_postojece(ws, config.HEADER_RED)
    ur = zadnji_ur + 1

    if k["match"]:
        red = k["match"]["red"]                       # postojeći UR0 red s uplatom
        for c in (2, 4, 7, 8, 10, 11):
            excel_mod.stil_celije(ws, red, c)
        ws.cell(row=red, column=1).value = ur
        c1 = ws.cell(row=red, column=1)
        c1.font = Font(name=c1.font.name, size=c1.font.size, color="000000", underline=None)
    else:
        zadnji_red, _, _ = excel_mod.analiziraj_postojece(ws, config.HEADER_RED)
        red = zadnji_red + 1
        excel_mod.upisi_vrijednosti(ws, red, {1: ur}, stil_izvor_red=red - 1)  # stil + UR + formule

    ws.cell(row=red, column=2).value = polja.get("broj")
    if polja.get("datum"):
        ws.cell(row=red, column=4).value = polja["datum"]
    if polja.get("vozilo"):
        ws.cell(row=red, column=7).value = polja["vozilo"]
    ws.cell(row=red, column=8).value = polja.get("vrsta")
    ws.cell(row=red, column=10).value = polja.get("osnovica")
    ws.cell(row=red, column=11).value = polja.get("pdv")
    if polja.get("dobavljac") and not ws.cell(row=red, column=5).value:
        ws.cell(row=red, column=5).value = polja["dobavljac"]

    # slika -> "UR XXXX.pdf" + link (samo u stvarnom radu / prisili)
    if premjesti:
        novo_ime = f"UR {str(ur).zfill(4)}.pdf"
        cilj = os.path.join(config.FOLDER_PDF_FINAL, novo_ime)
        try:
            fotke_mod.pretvori_u_pdf(k["slika"], cilj)
            excel_mod.postavi_hyperlink(ws, red, 1, cilj)
            os.remove(k["slika"])           # makni iz privremene mape (Desktop\URA)
            logger.info("   📄 fotka → %s (+ link)", novo_ime)
        except Exception as e:
            logger.error("   ⚠️ Ne mogu pretvoriti/povezati sliku: %s", e)

    if pisi:
        excel_mod.spremi(wb, config.EXCEL_PATH, logger)
        stanje_mod.spremi(stanje, config.STANJE_PATH)
    if k["match"]:
        try:
            sesija["uplate"].remove(k["match"])   # potrošena — da je iduća fotka ne uzme opet
        except ValueError:
            pass
    logger.info("✅ UR%04d upisan (red %s): %s, %s €", ur, red, polja.get("broj"), polja.get("osnovica"))
    return ur


def fale_podaci_sesija(logger):
    """Otvori knjigu i vrati račune kojima FALI VRSTA TROŠKA (ekran 'Fale podaci').
    Drži wb otvoren (kao fotke); svaki red dobije putanju do svog PDF-a (UR XXXX.pdf)."""
    wb = excel_mod.ucitaj_workbook(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    redovi = izvjestaji_mod.racuni_bez_vrste(ws, config.HEADER_RED, config.OBRADJUJ_OD)
    for r in redovi:
        r["pdf"] = os.path.join(config.FOLDER_PDF_FINAL, f"UR {str(r['ur']).zfill(4)}.pdf")
    excel_mod.pripremi_stil(ws, config.HEADER_RED)
    logger.info("FALE PODACI (vrsta troška) — %s računa", len(redovi))
    return {"ws": ws, "wb": wb, "redovi": redovi, "backup_done": False}


def upisi_vrstu(sesija, idx, vrsta, logger):
    """Upiši VRSTU TROŠKA (kolona 8) u zadani red i spremi knjigu (backup prije prvog pisanja)."""
    ws, wb = sesija["ws"], sesija["wb"]
    r = sesija["redovi"][idx]
    red = r["red"]
    if not config.SIMULACIJA and not sesija["backup_done"]:
        backup_mod.napravi_backup(config.EXCEL_PATH, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
        sesija["backup_done"] = True
    excel_mod.stil_celije(ws, red, 8)            # zadrži stil ćelije
    ws.cell(row=red, column=8).value = vrsta
    if not config.SIMULACIJA:
        excel_mod.spremi(wb, config.EXCEL_PATH, logger)
    logger.info("✅ UR%04d: vrsta troška = %s", r["ur"], vrsta)
    return r["ur"]


def zatvori_sesiju(sesija):
    """Zatvori otvorenu knjigu iz sesije (ekrani 'Fale podaci' / fotke / spoji uplate)."""
    try:
        sesija["wb"].close()
    except Exception:
        pass


def kandidati_spajanja(logger):
    """Nađi MOGUĆA spajanja: neplaćeni računi + 'viseće' uplate s izvoda (UR0), upareni
    po DOBAVLJAČU + IZNOSU (za slučajeve gdje spajanje po broju nije uspjelo). Vrati sesiju
    (drži knjigu otvorenu); upis tek na potvrdu (potvrdi_spajanje)."""
    from rapidfuzz import fuzz
    wb = excel_mod.ucitaj_workbook(config.EXCEL_PATH)
    ws = wb[config.GLAVNI_SHEET]
    excel_mod.pripremi_stil(ws, config.HEADER_RED)
    neplaceni = izvjestaji_mod.racuni_bez_izvoda(ws, config.HEADER_RED, None)
    uplate = izvjestaji_mod.izvodi_bez_racuna(ws, config.HEADER_RED, None)

    def _d(v):
        return v.date() if hasattr(v, "date") else None

    parovi, iskoristene = [], set()
    for inv in sorted(neplaceni, key=lambda r: str(r.get("datum"))):
        iv = inv.get("ukupno")
        if iv is None:
            continue
        kand = []
        for up in uplate:
            if up["red"] in iskoristene or up.get("placeno") is None:
                continue
            try:
                if abs(float(iv) - float(up["placeno"])) > 0.02:
                    continue
            except (TypeError, ValueError):
                continue
            sim = fuzz.partial_ratio(str(inv.get("dobavljac") or "").lower(),
                                     str(up.get("dobavljac") or "").lower())
            if sim < 80:
                continue
            di, dp = _d(inv.get("datum")), _d(up.get("datum"))
            dd = abs((di - dp).days) if (di and dp) else 999
            kand.append((sim, dd, up))
        if not kand:
            continue
        kand.sort(key=lambda x: (-x[0], x[1]))          # najsličnije ime, pa najbliži datum
        up = kand[0][2]
        iskoristene.add(up["red"])
        lk = ws.cell(row=up["red"], column=3).hyperlink
        parovi.append({
            "r_inv": inv["red"], "ur": inv["ur"], "broj": inv["racun"],
            "dob_inv": inv.get("dobavljac"), "dat_inv": inv.get("datum"), "iznos": iv,
            "r_pay": up["red"], "izvadak": up.get("izvadak"), "dob_pay": up.get("dobavljac"),
            "dat_pay": up.get("datum"), "placeno": up.get("placeno"),
            "sred": ws.cell(row=up["red"], column=18).value,
            "izvod_link": (lk.target if lk else None),
        })
    logger.info("Mogućih spajanja (dobavljač + iznos): %s", len(parovi))
    return {"wb": wb, "ws": ws, "parovi": parovi, "backup_done": False}


def potvrdi_spajanje(sesija, idx, logger):
    """Potvrđeno spajanje: prenesi uplatu (izvadak/datum/plaćeno/sredstvo + link) s 'viseće'
    UR0 stavke na račun, pa isprazni tu UR0 stavku (da se iznos ne broji dvaput). Spremi."""
    ws, wb = sesija["ws"], sesija["wb"]
    p = sesija["parovi"][idx]
    r_inv, r_pay = p["r_inv"], p["r_pay"]
    if ws.cell(row=r_pay, column=16).value in (None, ""):
        return {"greska": "Uplata je već iskorištena/prazna."}
    if not sesija["backup_done"]:
        backup_mod.napravi_backup(config.EXCEL_PATH, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
        sesija["backup_done"] = True
    for col in (3, 15, 16, 18):
        excel_mod.stil_celije(ws, r_inv, col)
    ws.cell(row=r_inv, column=3).value = ws.cell(row=r_pay, column=3).value   # IZVADAK
    ws.cell(row=r_inv, column=15).value = ws.cell(row=r_pay, column=15).value  # datum plać
    ws.cell(row=r_inv, column=16).value = ws.cell(row=r_pay, column=16).value  # PLAĆENO
    ws.cell(row=r_inv, column=18).value = ws.cell(row=r_pay, column=18).value  # sredstvo
    if p.get("izvod_link"):
        excel_mod.postavi_hyperlink(ws, r_inv, 3, p["izvod_link"])
    c3 = ws.cell(row=r_pay, column=3)
    if c3.hyperlink:
        c3.hyperlink = None
    for col in range(1, 21):                                                   # isprazni UR0 stavku
        ws.cell(row=r_pay, column=col).value = None
    excel_mod.spremi(wb, config.EXCEL_PATH, logger)
    logger.info("✅ Spojeno: UR%s ← izvod %s (%.2f €)", p["ur"], p["izvadak"], p["placeno"])
    return {"ok": True, "prazan_red": r_pay}


def _procitaj_tereni(godina, mjesec):
    """Pročitaj terene za (godina, mjesec) iz 'tereni' tablice (list 'tereni {godina}').
    Stupci: [1]polazak [2]povratak [3]ime [4]auto [5]km prije [6]km posl [7]lokacija [8]razlog."""
    import openpyxl
    if not os.path.exists(config.TERENI_PATH):
        return None
    wb = openpyxl.load_workbook(config.TERENI_PATH, data_only=True, read_only=True)
    sheet = f"tereni {godina}"
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        pol = r[1] if len(r) > 1 else None
        if not hasattr(pol, "month") or pol.year != godina or pol.month != mjesec:
            continue
        pov = r[2] if (len(r) > 2 and hasattr(r[2], "date")) else pol
        out.append({
            "djelatnik": (str(r[3]).strip() if r[3] else ""),
            "auto": (str(r[4]).strip() if r[4] else ""),
            "lokacija": (str(r[7]).strip() if (len(r) > 7 and r[7]) else ""),
            "razlog": (str(r[8]).strip() if (len(r) > 8 and r[8]) else ""),
            "polazak": pol.date(),
            "povratak": pov.date() if hasattr(pov, "date") else pol.date(),
        })
    wb.close()
    return out


def _sirocad_prolazaka(prolasci, pokriveno):
    """Vrati ENC prolaske čije vozilo taj dan NEMA svoj teren ('siročad' — vjerojatno
    posuđen uređaj, npr. pokvaren Jumpyjev pa je uzet Fordov). 'pokriveno' = lista
    (vozilo_norm, d1, d2) za SVE terene mjeseca."""
    return [p for p in prolasci
            if not any(v == p["voz"] and d1 <= p["datum"] <= d2 for v, d1, d2 in pokriveno)]


def generiraj_naloge(godina, mjesec, enc_csv, logger):
    """Izradi SVE putne naloge za mjesec iz 'tereni' tablice (+ ENC vremena). Vrati broj izrađenih."""
    logger.info("=" * 64)
    logger.info("GENERIRANJE NALOGA — %02d-%s", mjesec, godina)
    tereni = _procitaj_tereni(godina, mjesec)
    if tereni is None:
        logger.error("Ne nalazim 'tereni' tablicu: %s", config.TERENI_PATH)
        return 0
    logger.info("Terena u tablici za taj mjesec: %s", len(tereni))
    if not tereni:
        return 0
    pn_path, je_novi = _osiguraj_pn(godina, mjesec, logger)
    if not pn_path:
        return 0
    ima_enc = bool(enc_csv and os.path.exists(enc_csv))
    enc = vremena_mod.enc_grupe(enc_csv) if ima_enc else {}
    prolasci = vremena_mod.enc_prolasci(enc_csv) if ima_enc else []
    logger.info("ENC prolazaka učitano: %s%s", len(enc), "" if enc else "  (sva vremena standardna)")

    # dedup: postojeći nalozi u PN Excelu (djelatnik + datum + mjesto)
    import openpyxl
    wbv = openpyxl.load_workbook(pn_path, data_only=True)
    postojeci = set()
    for sh in wbv.sheetnames:
        if not sh.strip().isdigit():
            continue
        w = wbv[sh]
        dat = w["E6"].value
        d = dat.date() if hasattr(dat, "date") else dat
        postojeci.add((str(w["E4"].value or "").strip().lower(), str(d),
                       str(w["P4"].value or "").strip().lower()))
    wbv.close()

    lista = []
    meta = []          # teren za svaki element 'liste' (za pripis posuđenih ENC prolazaka)
    for t in tereni:
        kljuc = (t["djelatnik"].lower(), str(t["polazak"]), t["lokacija"].lower())
        if kljuc in postojeci:
            continue  # nalog već postoji
        puni, enc_nick = config.vozilo_info(t["auto"])
        od, na = vremena_mod.vremena(t["djelatnik"], enc_nick or "", t["polazak"], t["povratak"], enc)
        # ENC prolasci -> cestarina u specifikaciji troškova (hac, tvrtka)
        troskovi = []
        if enc_nick:
            for p in vremena_mod.prolasci_za_teren(prolasci, enc_nick, t["polazak"], t["povratak"]):
                troskovi.append({"opis": "cestarina", "izdavatelj": "hac",
                                 "datum": p["datum"], "iznos": p["iznos"], "sredstvo": "tvrtka"})
        # svrha = jedna od 3 opcije iz padajućeg (E5): obilazak gradilišta / sastanak / ispitivanje
        razlog = t["razlog"] or ""
        rl = razlog.lower()
        if "obilazak" in rl and "gradiliš" in rl:
            svrha = "obilazak gradilišta"
        elif "sastanak" in rl:
            svrha = "sastanak"
        else:
            svrha = "ispitivanje"
        lista.append({
            "djelatnik": t["djelatnik"], "mjesto": t["lokacija"],
            "svrha": svrha,            # padajući izbornik (E5)
            "izvjesce": razlog,        # puni tekst razloga -> kućica 'Izvješće s puta' (A34)
            "datum_odlaska": od.date(), "vrijeme_odlaska": od.time(),
            "datum_povratka": na.date(), "vrijeme_povratka": na.time(),
            "vozilo": puni, "vrsta_prijevoza": "službeno",
            "drzava": "HR", "dnevnica_iznos": 30, "locco": 0.4,
            "troskovi": troskovi,
        })
        meta.append(t)

    # POSUĐENI ENC UREĐAJ: kad nekome ne radi uređaj pa uzme tuđi (npr. Jumpy vozi s
    # Fordovim), prolaz je zabilježen pod krivim vozilom. Takve prolaske ("siročad" —
    # vozilo bez terena taj dan) pripiši nalogu koji taj dan NEMA nijedan svoj prolaz.
    # Samo ako je kandidat JEDNOZNAČAN — inače upozorenje pa korisnica upiše ručno.
    pokriveno = []
    for t in tereni:
        _, nk = config.vozilo_info(t["auto"])
        if nk:
            pokriveno.append((vremena_mod._voz(nk), t["polazak"], t["povratak"]))
    grupe = {}
    for p in _sirocad_prolazaka(prolasci, pokriveno):
        grupe.setdefault((p["voz"], p["datum"]), []).append(p)
    for (voz, dan), grupa in sorted(grupe.items()):
        kand = [i for i, n in enumerate(lista)
                if not n["troskovi"] and meta[i]["polazak"] <= dan <= meta[i]["povratak"]]
        if len(kand) == 1:
            i = kand[0]
            for p in grupa:
                lista[i]["troskovi"].append({"opis": "cestarina", "izdavatelj": "hac",
                                             "datum": p["datum"], "iznos": p["iznos"],
                                             "sredstvo": "tvrtka"})
            logger.info("↔️ Posuđeni ENC uređaj '%s' (%s): %s prolaza → nalog %s / %s",
                        voz, dan, len(grupa), meta[i]["djelatnik"], meta[i]["auto"])
        elif kand:
            logger.warning("⚠️ ENC prolasci uređaja '%s' (%s) nemaju svoj teren, a %s je "
                           "mogućih naloga — upiši ručno.", voz, dan, len(kand))
        else:
            logger.warning("⚠️ ENC prolasci uređaja '%s' (%s) nemaju svoj teren ni nalog "
                           "bez prolazaka — provjeri/upiši ručno.", voz, dan)

    if not lista:
        logger.info("Svi tereni već imaju nalog — ništa novo.")
        return 0
    logger.info("Za izradu: %s naloga (preskočeno postojećih: %s)", len(lista), len(tereni) - len(lista))
    try:
        backup_mod.napravi_backup(pn_path, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
    except Exception as e:
        logger.error("Backup PN Excela nije uspio (%s). Prekidam.", e)
        return 0
    brojevi = putni_izrada_mod.kreiraj_naloge_batch(pn_path, lista, logger, ocisti_stare=je_novi)
    if brojevi:
        try:
            os.startfile(pn_path)   # otvori PN Excel (aktivan je najmanji nalog)
        except Exception:
            pass
    return len(brojevi)


def _osiguraj_pn(godina, mjesec, logger):
    """Vrati (put_za_pisanje, je_novi_mjesec).
    - U test načinu piše u KOPIJU (test_data/PN_test); pravi PN Excel se NE dira.
    - Ako mjesečni PN Excel NE postoji → napravi ga iz PRETHODNOG mjeseca (je_novi=True;
      tada generator poslije makne prenesene stare naloge, ostaju samo novi)."""
    cilj = config.pn_excel_putanja(godina, mjesec)
    pravi = config.pn_excel_pravi(godina, mjesec)
    test_mode = config.FOLDER_PN != config.FOLDER_PN_PRAVI

    # 1) mjesec POSTOJI (na serveru): u testu napravi kopiju, vrati; nije novi
    if os.path.exists(pravi):
        if test_mode and not os.path.exists(cilj):
            os.makedirs(os.path.dirname(cilj), exist_ok=True)
            shutil.copy2(pravi, cilj)
            logger.info("TEST: kopija PN Excela → %s", cilj)
        return (cilj if os.path.exists(cilj) else None), False
    if test_mode and os.path.exists(cilj):
        return cilj, False   # već postoji test-kopija tog mjeseca

    # 2) mjesec NE postoji → napravi iz najbližeg ranijeg mjeseca
    prethodni = None
    for m in range(mjesec - 1, 0, -1):
        pp = config.pn_excel_pravi(godina, m)
        if os.path.exists(pp):
            prethodni = pp
            break
    if not prethodni:
        logger.error("Nema nijednog ranijeg PN Excela u %s za osnovu novog mjeseca.", godina)
        return None, False
    os.makedirs(os.path.dirname(cilj), exist_ok=True)
    shutil.copy2(prethodni, cilj)
    logger.info("Napravljen NOVI mjesečni PN Excel %s (osnova: %s)",
                os.path.basename(cilj), os.path.basename(prethodni))
    return cilj, True


def putni_popisi(godina, mjesec):
    """Padajući popisi (djelatnici/auti/svrhe/vrste/države) za formular izrade naloga.
    Čita iz PRAVOG mjesečnog PN Excela; ako tog nema, proba zadnji dostupni mjesec u godini."""
    p = config.pn_excel_pravi(godina, mjesec)
    if not os.path.exists(p):
        for m in range(12, 0, -1):
            alt = config.pn_excel_pravi(godina, m)
            if os.path.exists(alt):
                p = alt
                break
    if not os.path.exists(p):
        return None
    return putni_izrada_mod.popisi(p)


def kreiraj_putni(godina, mjesec, podaci, logger):
    """Izradi novi putni nalog u mjesečnom PN Excelu (preko Excela, uz backup). Vrati broj."""
    p, _ = _osiguraj_pn(godina, mjesec, logger)
    if not p:
        return None
    try:
        backup_mod.napravi_backup(p, config.BACKUP_DIR, config.BACKUP_KEEP, logger)
    except Exception as e:
        logger.error("Backup PN Excela nije uspio (%s). Prekidam.", e)
        return None
    return putni_izrada_mod.kreiraj_nalog(p, podaci, logger)


def brojke_za_plocice(od_datuma=None):
    """Brojevi za pločice na vrhu aplikacije (samo čita knjigu)."""
    try:
        wb = excel_mod.ucitaj_workbook_vrijednosti(config.EXCEL_PATH)
        ws = wb[config.GLAVNI_SHEET]
        rez = izvjestaji_mod.brojke(ws, config.HEADER_RED, od_datuma)
        wb.close()
        return rez
    except Exception:
        return {"racuni": 0, "ceka_izvod": 0, "treba_potvrdu": 0, "za_rucno": 0}


def obrisi_pamcenje(logger):
    """Očisti zapamćeno stanje (stanje.json) — za ponovni čisti test. Radi backup prije."""
    import shutil
    p = config.STANJE_PATH
    if os.path.exists(p):
        bkp = p + ".bak"
        try:
            shutil.copy2(p, bkp)
        except Exception:
            pass
        os.remove(p)
        logger.info("🧹 Pamćenje očišćeno (backup: %s). Sljedeća obrada kreće ispočetka.",
                    os.path.basename(bkp))
        return True
    logger.info("Nema zapamćenog stanja za brisanje.")
    return False


def pokreni():
    """Konzolni ulaz: izbornik 1/2 pa jezgra obrade."""
    logger = postavi_logging(config.LOG_DIR)
    nacin = izbornik()
    pokreni_obradu(nacin, logger)


def _izvjestaj(logger, stat):
    logger.info("=" * 64)
    logger.info("IZVJEŠTAJ:")
    logger.info("  Skinuto s Parre:          %s", stat["skinuto"])
    logger.info("  Već u knjizi:             %s (po broju %s + duplikat %s)",
                stat["vec_po_broju"] + stat["duplikat"], stat["vec_po_broju"], stat["duplikat"])
    logger.info("  Novi za upis:             %s (od toga 'mogući duplikat' %s) → upisano %s",
                stat["novo"] + stat["neizvjesno"], stat["neizvjesno"], stat["upisano"])
    logger.info("  Deferno (dopunjena uplata): %s", stat["deferno"])
    logger.info("  PDF povezan:              %s", stat["pdf_spojen"])
    logger.info("  PDF nije nađen:           %s", stat["pdf_nije_nadjen"])
    logger.info("  Datoteke → Parra mapu:    %s", stat["parra_premjesteno"])
    logger.info("  Poveznice dopunjene:      %s", stat["linkovi_dodani"])
    logger.info("  --- IZVODI ---")
    logger.info("  Izvoda obrađeno:          %s", stat["izvoda_obradeno"])
    logger.info("  Spojeno s računom:        %s", stat["spojeno"])
    logger.info("  Dnevnice spojene:         %s", stat["dnevnice"])
    logger.info("  Novih UR0 redova:         %s", stat["ur0_redova"])
    logger.info("  Uplate preskočene:        %s", stat["preskoceno_uplata"])
    logger.info("  Greške:                   %s", stat["greske"])
    logger.info("=" * 64)


if __name__ == "__main__":
    try:
        pokreni()
    except Exception:
        import logging
        logging.getLogger("ura").error("NEOČEKIVANA GREŠKA:\n%s", traceback.format_exc())
        traceback.print_exc()
    finally:
        try:
            input("\nPritisni Enter za kraj...")
        except EOFError:
            pass
