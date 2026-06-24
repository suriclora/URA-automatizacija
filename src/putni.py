# -*- coding: utf-8 -*-
"""
FAZA 2 — PUTNI NALOZI.

korisnik izrađuje putne naloge u mjesečnom Excelu na serveru
(PN/{godina}/pn MM-GGGG.xlsx), gdje je SVAKI LIST jedan nalog (npr. '033').
Umjesto OCR-a skeniranih PDF-ova, čitamo podatke IZRAVNO iz tog Excela —
uredno i bez grešaka prepoznavanja.

Iz svakog lista vadimo: broj naloga, djelatnik, datum odlaska, iznos za isplatu.
"""
from datetime import datetime

import openpyxl


def _val_desno(ws, *labeli):
    """Vrati prvu ne-praznu vrijednost DESNO od ćelije čiji tekst sadrži neki od labela."""
    for row in ws.iter_rows():
        for idx, c in enumerate(row):
            if isinstance(c.value, str) and any(l in c.value.lower() for l in labeli):
                for c2 in row[idx + 1:]:
                    if c2.value not in (None, ""):
                        return c2.value
    return None


def _datum(v):
    if isinstance(v, datetime):
        return v.date()
    return None


def procitaj_naloge(excel_path):
    """Pročitaj sve naloge iz mjesečnog PN Excela. Jedan list = jedan nalog.
    Vrati listu rječnika: {broj, djelatnik, datum, iznos}.
    Preskoči listove bez djelatnika ili bez iznosa (prazni/nedovršeni nalozi)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    out = []
    for sh in wb.sheetnames:
        ws = wb[sh]
        broj = _val_desno(ws, "putni nalog broj")
        djelatnik = _val_desno(ws, "djelatnik:")
        datum = _datum(_val_desno(ws, "datum odlaska"))
        iznos = _val_desno(ws, "za isplatu")
        if not djelatnik or iznos in (None, "", 0):
            continue  # prazan ili nedovršen nalog
        try:
            iznos = float(iznos)
        except (TypeError, ValueError):
            continue
        out.append({
            "broj": str(broj or sh).strip().lstrip("0").zfill(3),  # npr. '033'
            "djelatnik": str(djelatnik).strip(),
            "datum": datum,
            "iznos": iznos,
        })
    wb.close()
    return out
