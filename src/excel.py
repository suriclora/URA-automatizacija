# -*- coding: utf-8 -*-
"""
Sav rad s Excelom: učitavanje, analiza postojećih redova, upis i —
najvažnije — ČUVANJE FORMATIRANJA (font, boje, okviri, plavi hiperlinkovi).

Funkcije za stil su preuzete IDENTIČNO iz postojeće skripte, jer su radile
dobro i korisnik traži da formatiranje ostane nepromijenjeno.
"""
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font


# ---------------------------------------------------------------------
#  STILOVI  (preuzeto iz originalne skripte — ne diramo logiku)
# ---------------------------------------------------------------------

def kopiraj_stil_identicno(izvorna, nova):
    """Prekopira izgled (font, okvir, ispunu, format broja, poravnanje)
    iz jedne ćelije u drugu — tako da novi red izgleda kao postojeći."""
    if izvorna.has_style:
        nova.font = copy(izvorna.font)
        nova.border = copy(izvorna.border)
        nova.fill = copy(izvorna.fill)
        nova.number_format = copy(izvorna.number_format)
        nova.alignment = copy(izvorna.alignment)


def postavi_link_stil(celija):
    """Postavi stil hiperlinka: plavo, BEZ podcrtavanja (kao postojeći redovi)."""
    stari = celija.font
    celija.font = Font(
        name=stari.name,
        size=stari.size,
        bold=stari.bold,
        color="0000FF",
        underline=None,
    )


# ---------------------------------------------------------------------
#  UČITAVANJE I ANALIZA
# ---------------------------------------------------------------------

def ucitaj_workbook(path):
    """Učitaj knjigu za PISANJE. .xlsm -> keep_vba=True (čuva makroe);
    .xlsx (npr. demo) -> bez makroa."""
    keep = str(path).lower().endswith(".xlsm")
    return load_workbook(path, keep_vba=keep)


def ucitaj_workbook_vrijednosti(path):
    """Učitaj za ČITANJE izračunatih vrijednosti formula (data_only=True).
    NE read_only — jer koristimo i ws.cell() izravno (indeks spajanja, povijest vrste).
    """
    return load_workbook(path, data_only=True, read_only=False, keep_vba=False)


def analiziraj_postojece(ws, header_red):
    """Prođi list i vrati:
      - zadnji_red: redni broj zadnjeg reda s podacima
      - zadnji_ur:  najveći UR broj koji već postoji
      - postojeci:  skup već upisanih brojeva računa (mala slova) za dedup

    Robusnije od originala: ide do stvarnog zadnjeg reda (ws.max_row),
    bez fiksnog "range(2, 5000)".
    """
    zadnji_red = header_red
    zadnji_ur = 0
    postojeci = set()

    for i in range(header_red + 1, ws.max_row + 1):
        ur = ws.cell(row=i, column=1).value      # kolona A = UR
        rb = ws.cell(row=i, column=2).value      # kolona B = RAČUN BR
        if (ur is None or ur == "") and (rb is None or rb == ""):
            continue  # prazan red — preskoči

        zadnji_red = i
        if rb is not None and rb != "":
            postojeci.add(str(rb).strip().lower())
        try:
            broj = int(str(ur).split(".")[0])
            if broj > zadnji_ur:
                zadnji_ur = broj
        except (ValueError, TypeError):
            pass  # npr. "0000" ili prazno — preskoči

    return zadnji_red, zadnji_ur, postojeci


# ---------------------------------------------------------------------
#  UPIS
# ---------------------------------------------------------------------

MAX_STIL_KOL = 18  # do koje kolone kopiramo stil (UR..SREDSTVO)
UKUPNO_KOL = 12    # L =K+J
DUGOVANJE_KOL = 17 # Q =L-P

# Predložak stila PO STUPCU — izgrađen iz postojećih (ručnih) redova jednom na
# početku. Ključ je broj stupca, vrijednost je (font, border, fill, number_format,
# alignment). Koristimo ga umjesto kopiranja iz reda iznad, jer PRAZNE ćelije reda
# iznad nemaju stil (has_style=False) pa bi se kopija preskočila i ostao bi
# Excelov default (Arial 10 / General).
_STIL_PRED = None


def pripremi_stil(ws, header_red, max_col=MAX_STIL_KOL):
    """Izgradi predložak stila po stupcu IZ POSTOJEĆIH redova: za svaki stupac uzmi
    PRVU ćeliju koja IMA vrijednost (dakle je stilizirana). Pozvati JEDNOM, prije upisa."""
    global _STIL_PRED
    pred = {}
    for col in range(1, max_col + 1):
        for i in range(header_red + 1, ws.max_row + 1):
            c = ws.cell(row=i, column=col)
            if c.value not in (None, "") and c.has_style:
                pred[col] = (copy(c.font), copy(c.border), copy(c.fill),
                             copy(c.number_format), copy(c.alignment))
                break
    _STIL_PRED = pred
    return pred


def _primijeni_stupac(cell, stil):
    f, b, fl, nf, al = stil
    cell.font = copy(f)
    cell.border = copy(b)
    cell.fill = copy(fl)
    cell.number_format = copy(nf)
    cell.alignment = copy(al)


def stil_celije(ws, red, col):
    """Primijeni predložak stila na JEDNU ćeliju (npr. kad izvod puni PLAĆENO na
    postojećem ručnom retku koji je imao praznu, nestiliziranu ćeliju)."""
    if _STIL_PRED and col in _STIL_PRED:
        _primijeni_stupac(ws.cell(row=red, column=col), _STIL_PRED[col])


def kopiraj_red_stil(ws, src_red, dst_red, max_col=MAX_STIL_KOL):
    """Prekopira stil cijelog reda po stupcu. Ako postoji predložak (_STIL_PRED),
    koristi NJEGA (robusno); inače kopira iz zadanog reda (fallback)."""
    for col in range(1, max_col + 1):
        if _STIL_PRED and col in _STIL_PRED:
            _primijeni_stupac(ws.cell(row=dst_red, column=col), _STIL_PRED[col])
        else:
            kopiraj_stil_identicno(ws.cell(row=src_red, column=col),
                                   ws.cell(row=dst_red, column=col))


def upisi_vrijednosti(ws, red, podaci, stil_izvor_red):
    """Upiši vrijednosti u zadani red uz IDENTIČAN stil kao postojeći redovi
    (svaki stupac iz predloška). 'podaci' = {broj_kolone: vrijednost}.
    UKUPNO (L) i DUGOVANJE (Q) UVIJEK su formule (=K+J, =L-P), kao u knjizi.
    """
    kopiraj_red_stil(ws, stil_izvor_red, red)   # cijeli red — identičan izgled
    for col, val in podaci.items():
        if col in (UKUPNO_KOL, DUGOVANJE_KOL):
            continue  # formule postavljamo posebno, ispod
        ws.cell(row=red, column=col).value = val
    # UR ćelija: crni font (poslije, ako nađemo PDF, postaje plavi link)
    c1 = ws.cell(row=red, column=1)
    c1.font = Font(name=c1.font.name, size=c1.font.size, color="000000", underline=None)
    # Formule kao u knjizi
    ws.cell(row=red, column=UKUPNO_KOL).value = f"=K{red}+J{red}"
    ws.cell(row=red, column=DUGOVANJE_KOL).value = f"=L{red}-P{red}"


def postavi_hyperlink(ws, red, col, putanja):
    """Pretvori ćeliju u klikabilni link na PDF, sa stilom plavog linka."""
    cell = ws.cell(row=red, column=col)
    cell.hyperlink = putanja
    postavi_link_stil(cell)


def spremi(wb, path, logger):
    """Spremi workbook. Pozvati TEK NAKON što je napravljen backup."""
    wb.save(path)
    logger.info("Excel spremljen: %s", path)
