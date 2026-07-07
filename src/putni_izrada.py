# -*- coding: utf-8 -*-
"""
IZRADA putnih naloga — dodaje novi list u mjesečni PN Excel.

VAŽNO: piše se preko PRAVOG Excela (xlwings), NE openpyxl — jer openpyxl izbacuje
EMF logo i crteže s naloga (provjereno). xlwings koristi sam Excel pa logo, formule
i padajući popisi ostaju 100% netaknuti.

Broj naloga = IME LISTA (predložak ima E3 = MID(CELL("Filename"...)) → čita ime lista).
Sve dnevnice/za isplatu računaju formule u predlošku — mi punimo samo ulazna polja.
"""
import datetime as _dt
import os
import shutil

import openpyxl

import config


def _ocisti_genpy():
    """Počisti win32com 'gen_py' cache (zna se pokvariti od prekinutih Excel sesija)."""
    try:
        shutil.rmtree(os.path.join(os.environ.get("LOCALAPPDATA", ""), "Temp", "gen_py"),
                      ignore_errors=True)
    except Exception:
        pass


def _zatvori_app(app):
    """Pouzdano zatvori Excel (quit pa kill ako treba) — da ne ostanu zombi procesi."""
    try:
        app.quit()
    except Exception:
        try:
            app.kill()
        except Exception:
            pass


def _otvori_za_izradu(app, excel_path):
    """Otvori PN za izradu i vrati (wb, build_put).

    Excel-COM zna TRAJNO odbiti otvaranje baš imena 'pn MM-YYYY.xlsx' (čudna Excel
    pojava, neovisna o sadržaju i procesima — provjereno: identični bajtovi pod drugim
    imenom se otvore bez problema). Zato radimo na privremenoj kopiji U ISTOJ MAPI, a
    pozivatelj na kraju zove _spremi_natrag() koji je preimenuje natrag NA DISKU
    (Excel tako nikad ne mora COM-om otvoriti 'ukleto' ime)."""
    build = os.path.join(os.path.dirname(excel_path), "~izrada_pn.xlsx")
    try:
        if os.path.exists(build):
            os.remove(build)
    except Exception:
        pass
    shutil.copy2(excel_path, build)
    try:
        wb = app.books.open(build)
    except Exception:
        _ocisti_genpy()   # pokvareni COM cache se zna desiti -> počisti za idući put
        try:
            os.remove(build)
        except Exception:
            pass
        raise RuntimeError("Ne mogu otvoriti PN Excel za izradu. Pokušaj ponovo; "
                           "ako se ponavlja, javi.")
    return wb, build


def _spremi_natrag(build, excel_path):
    """Preimenuj privremenu kopiju natrag na pravo ime (na disku). Ako je pravi PN
    otvoren u Excelu, preimenovanje pada -> jasna poruka da ga zatvori."""
    try:
        os.replace(build, excel_path)
    except PermissionError:
        raise RuntimeError(f"Datoteka '{os.path.basename(excel_path)}' je otvorena u Excelu. "
                           "Zatvori je pa pokušaj ponovo.")


def _com(v):
    """Pretvori Python tip u nešto što Excel/COM prima (time->tekst, date->datetime)."""
    if isinstance(v, _dt.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, _dt.datetime):
        return v
    if isinstance(v, _dt.date):
        return _dt.datetime(v.year, v.month, v.day)
    return v

# Ulazna polja (ćelija -> ključ u 'podaci'); ostalo su formule koje NE diramo.
POLJA = {
    "E4": "djelatnik", "P4": "mjesto", "E5": "svrha",
    "E6": "datum_odlaska", "P6": "vrijeme_odlaska",
    "E7": "vozilo", "P7": "vrsta_prijevoza", "E8": "poc_brojilo",
    "E16": "datum_povratka", "P16": "vrijeme_povratka", "E18": "zav_brojilo",
    "E19": "drzava", "G19": "dnevnica_iznos", "P19": "locco",
    "A34": "izvjesce",   # tekst 'Izvješće s puta' (labela je u A33) — upisuje se nakon otključavanja
}


def popisi(excel_path):
    """Pročitaj padajuće popise iz predloška (djelatnici/auti/svrhe/vrste/države)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def stupac(col, r1=5, r2=25):
        out = []
        for r in range(r1, r2 + 1):
            v = ws[f"{col}{r}"].value
            s = str(v).strip() if v not in (None, "") else ""
            if s and not s.replace(".", "").replace(",", "").isdigit():  # preskoči brojeve (npr. stray '0')
                out.append(s)
        return out

    drzave = []
    for r in range(5, 12):
        d = ws[f"AH{r}"].value
        if d not in (None, ""):
            drzave.append((str(d).strip(), ws[f"AI{r}"].value))
    rez = {
        "djelatnici": stupac("Y"),
        "auti": stupac("AB"),
        "svrhe": stupac("AD"),
        "vrste": stupac("AF") or ["službeno", "privatno"],
        "drzave": drzave or [("HR", 30)],
    }
    wb.close()
    return rez


def sljedeci_broj(excel_path):
    """Najveći numerirani list + 1 (npr. '060')."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    brojevi = [int(s) for s in wb.sheetnames if s.strip().isdigit()]
    wb.close()
    return brojevi, (max(brojevi) + 1 if brojevi else 1)


def kreiraj_nalog(excel_path, podaci, logger):
    """Dodaj novi putni nalog (list) u mjesečni PN Excel preko Excela (xlwings).
    Vrati ime/broj novog naloga (npr. '060')."""
    import xlwings as xw

    brojevi, novi = sljedeci_broj(excel_path)
    ime = f"{novi:03d}"
    predlozak = f"{max(brojevi):03d}" if brojevi else None
    if not predlozak:
        raise RuntimeError("Nema postojećeg naloga za predložak u tom mjesecu.")

    build = None
    app = xw.App(visible=False)
    app.display_alerts = False
    try:
        wb, build = _otvori_za_izradu(app, excel_path)
        nov = _kopiraj_list(wb, predlozak, ime)
        _popuni_list(nov, podaci)
        wb.save()
        wb.close()
    finally:
        _zatvori_app(app)
    _spremi_natrag(build, excel_path)   # privremena kopija -> pravo ime (na disku)
    logger.info("✅ Putni nalog %s izrađen u %s", ime, excel_path)
    return ime


def _kopiraj_list(wb, predlozak, novo_ime):
    """Kopiraj list 'predlozak', otključaj ga i preimenuj u 'novo_ime'. Vrati novi list.
    Novu kopiju nalazimo po RAZLICI imena (pouzdanije od .active u petlji)."""
    sh = wb.sheets[predlozak]
    poslj = list(wb.sheets)[-1]               # zadnji list -> kopija ide NA KRAJ (rastući redoslijed)
    prije = [s.name for s in wb.sheets]
    sh.api.Copy(After=poslj.api)
    nov = next(s for s in wb.sheets if s.name not in prije)
    # listovi su zaštićeni LOZINKOM; ako je dana (config.PN_LOZINKA) otključaj posve
    try:
        nov.api.Unprotect(Password=getattr(config, "PN_LOZINKA", "") or "")
    except Exception:
        pass   # nema/ne valja lozinka -> pišemo samo u otključane (ulazne) ćelije
    nov.name = novo_ime
    return nov


def _popuni_list(nov, podaci):
    """Očisti ulazna polja + tablicu troškova i upiši podatke. Zaključane ćelije
    (npr. E19 država, troškovi — ako list nije otključan lozinkom) tiho preskačemo."""
    def pisi(adr, val):
        try:
            nov.range(adr).value = val
        except Exception:
            pass   # zaključana ćelija (zaštita) — preskoči
    for adr in POLJA:
        pisi(adr, None)
    # očisti LIJEVU (vidljivu) tablicu specifikacije ĆELIJU-PO-ĆELIJU. Čišćenje cijelog
    # raspona (clear_contents na A42:P48) zna TIHO pasti na zaštićenim/spojenim ćelijama —
    # tada u novom nalogu ostanu cestarine iz predloška (krivi datumi/iznosi)!
    for _r in range(42, 49):
        for _c in ("A", "D", "G", "K", "M", "P"):
            pisi(f"{_c}{_r}", None)
    for adr, kljuc in POLJA.items():
        v = podaci.get(kljuc)
        if v not in (None, ""):
            pisi(adr, _com(v))
    # cestarine -> LIJEVA specifikacija: A=opis, D=račun(prazno), G=izdavatelj, K=datum, M=iznos, P=sredstvo
    for i, t in enumerate((podaci.get("troskovi") or [])[:7]):
        r = 42 + i
        pisi(f"A{r}", t.get("opis"))
        pisi(f"G{r}", t.get("izdavatelj"))
        if t.get("datum") is not None:
            pisi(f"K{r}", _com(t.get("datum")))
        if t.get("iznos") is not None:
            pisi(f"M{r}", t.get("iznos"))
        pisi(f"P{r}", t.get("sredstvo"))


def kreiraj_naloge_batch(excel_path, lista_podataka, logger, ocisti_stare=False):
    """Izradi VIŠE naloga odjednom u jednoj Excel sesiji (brže). 'lista_podataka' = lista
    rječnika. ocisti_stare=True (novi mjesec): nakon izrade obriši PRENESENE stare naloge
    (iz prethodnog mjeseca), ostaju samo novi. Vrati listu brojeva izrađenih naloga."""
    import xlwings as xw

    brojevi, _ = sljedeci_broj(excel_path)
    if not brojevi:
        raise RuntimeError("Nema postojećeg naloga za predložak u tom mjesecu.")
    predlozak = f"{max(brojevi):03d}"
    sljedeci = max(brojevi) + 1
    stvoreni = []
    build = None
    app = xw.App(visible=False)
    app.display_alerts = False
    try:
        try:
            app.api.ScreenUpdating = False
            app.api.Calculation = -4135   # xlCalculationManual (Excel ne preračunava nakon svakog upisa)
        except Exception:
            pass
        wb, build = _otvori_za_izradu(app, excel_path)
        stari = [s.name for s in wb.sheets if s.name.strip().isdigit()] if ocisti_stare else []
        for podaci in lista_podataka:
            ime = f"{sljedeci:03d}"
            nov = _kopiraj_list(wb, predlozak, ime)
            _popuni_list(nov, podaci)
            stvoreni.append(ime)
            sljedeci += 1
        try:
            app.api.Calculation = -4105   # xlCalculationAutomatic (preračunaj dnevnice prije spremanja)
        except Exception:
            pass
        if ocisti_stare and stvoreni:
            for naziv in stari:                       # makni naloge prenesene iz prethodnog mjeseca
                try:
                    wb.sheets[naziv].api.Delete()
                except Exception:
                    pass
            logger.info("Maknuto %s prenesenih (starih) naloga iz novog mjeseca.", len(stari))
        # aktiviraj NAJMANJI nalog (da Excel otvori baš na njemu)
        try:
            numerirani = sorted((s for s in wb.sheets if s.name.strip().isdigit()),
                                key=lambda s: int(s.name))
            if numerirani:
                numerirani[0].activate()
        except Exception:
            pass
        wb.save()
        wb.close()
    finally:
        _zatvori_app(app)
    _spremi_natrag(build, excel_path)   # privremena kopija -> pravo ime (na disku)
    logger.info("✅ Izrađeno %s naloga u %s", len(stvoreni), excel_path)
    return stvoreni
