# -*- coding: utf-8 -*-
"""
Generator IZMIŠLJENOG demo skupa podataka.

Pokretanje:  python demo/generiraj_demo.py

Stvara `demo/data/` s potpuno lažnim podacima (nijedan stvaran podatak):
  - URA_demo.xlsx     — knjiga ulaznih računa (list 'UR'), miks stanja
  - tereni_demo.xlsx  — tablica terena (za izradu putnih naloga)
  - parra_demo.json   — lažni "ulazni e-računi" (zamjena za Parra API u DEMO načinu)
  - prazne podmape:    pdf/ Parra/ Downloads/ Izvodi/{HPB,RBA}/ PN/ fotke/ backup/

Sve je nasumično, ali s fiksnim sjemenom (reproducibilno).
"""
import os
import json
import random
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl

random.seed(42)
BASE = Path(__file__).resolve().parent / "data"

DOBAVLJACI = [
    ("ALFA TRGOVINA d.o.o.", "11111111111"),
    ("BETA SERVIS d.o.o.", "22222222222"),
    ("GAMA ENERGIJA d.o.o.", "33333333333"),
    ("DELTA LEASING d.d.", "44444444444"),
    ("EPSILON TELEKOM d.d.", "55555555555"),
    ("ZETA GRADNJA d.o.o.", "66666666666"),
    ("ETA OSIGURANJE d.d.", "77777777777"),
    ("THETA AUTOCESTE d.o.o.", "88888888888"),
]
DJELATNICI = ["Pero Perić", "Ivan Ivić", "Marko Marić", "Ana Anić"]
VOZILA = ["Kombi 1 (teretni)", "Osobno 1", "Dostavno 1 (teretni)"]
SREDSTVA = ["virman", "kartica", "gotovina"]


def _dt(y, m, d):
    return datetime(y, m, d)


def stvori_mape():
    for sub in ["", "pdf", "Parra", "Downloads", "Izvodi/HPB", "Izvodi/RBA",
                "PN/2026", "fotke", "backup"]:
        (BASE / sub).mkdir(parents=True, exist_ok=True)


def knjiga_ura():
    """URA_demo.xlsx — list 'UR', zaglavlje u redu 2, podaci od reda 3."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UR"
    zagl = ["UR", "RAČUN BR", "IZVADAK BR", "DATUM", "DOBAVLJAČ", "DJELATNIK",
            "VOZILO", "VRSTA TROŠKA", "NA TERET", "TROŠAK", "PDV", "UKUPNO",
            "ROK PL", "VAL", "DATUM PL", "PLAĆENO", "DUGOVANJE", "SREDSTVO",
            "BROJ KARTICE", "info"]
    for j, h in enumerate(zagl, start=1):
        ws.cell(row=2, column=j, value=h)

    r = 3
    ur = 100

    def red(ur_v, racun, izvadak, datum, dob, vrsta, neto, pdv, placeno, sredstvo, djel=""):
        nonlocal r
        ws.cell(row=r, column=1, value=ur_v)
        ws.cell(row=r, column=2, value=racun)
        ws.cell(row=r, column=3, value=izvadak)
        ws.cell(row=r, column=4, value=datum)
        ws.cell(row=r, column=5, value=dob)
        ws.cell(row=r, column=6, value=djel)
        ws.cell(row=r, column=8, value=vrsta)
        ws.cell(row=r, column=10, value=neto)
        ws.cell(row=r, column=11, value=pdv)
        ws.cell(row=r, column=12, value=f"=K{r}+J{r}")     # UKUPNO formula (kao u pravoj knjizi)
        ws.cell(row=r, column=16, value=placeno)
        ws.cell(row=r, column=17, value=f"=L{r}-P{r}")     # DUGOVANJE formula
        ws.cell(row=r, column=18, value=sredstvo)
        r += 1

    # 1) RAČUNI bez izvoda (UR>0, ima broj i vrstu, NIJE plaćeno) -> "fali izvod"
    for i in range(10):
        ur += 1
        dob, _ = random.choice(DOBAVLJACI)
        neto = round(random.uniform(40, 1800), 2)
        pdv = round(neto * 0.25, 2)
        red(ur, f"{random.randint(100,9999)}/1/1", None, _dt(2026, 5, random.randint(2, 28)),
            dob, random.choice(["gorivo", "uredske potrepštine", "električna energija",
                                 "knjigovodstvo", "cestarina"]),
            neto, pdv, None, "virman")

    # 2) RAČUNI plaćeni (ima izvadak + plaćeno) -> samo broje se kao upisani
    for i in range(6):
        ur += 1
        dob, _ = random.choice(DOBAVLJACI)
        neto = round(random.uniform(50, 900), 2)
        pdv = round(neto * 0.25, 2)
        red(ur, f"{random.randint(100,9999)}/1/1", random.randint(40, 95),
            _dt(2026, 5, random.randint(2, 28)), dob,
            random.choice(["gorivo", "leasing", "fiksna usluga"]),
            neto, pdv, round(neto + pdv, 2), "virman")

    # 3) RAČUNI bez VRSTE (skripta nije bila sigurna) -> "treba potvrdu"
    for i in range(3):
        ur += 1
        dob, _ = random.choice(DOBAVLJACI)
        neto = round(random.uniform(30, 400), 2)
        red(ur, f"{random.randint(100,9999)}/1/1", None, _dt(2026, 6, random.randint(2, 20)),
            dob, None, neto, round(neto * 0.25, 2), None, "virman")

    # 4) UPLATE s izvoda bez računa (UR 0000, plaćeno, vrsta NIJE trajna) -> "izvodima fale računi"
    for i in range(4):
        dob, _ = random.choice(DOBAVLJACI)
        izn = round(random.uniform(20, 300), 2)
        red(0, None, random.randint(40, 95), _dt(2026, 5, random.randint(2, 28)),
            dob, random.choice(["gorivo", "uredske potrepštine"]), izn, 0, izn, "kartica")

    # 5) Trajni UR 0000 (plaća/porez/dnevnice) -> NE ulaze u "fale računi"
    for vrsta, izn in [("plaća", 1850.0), ("porez", 640.0), ("dnevnice", 200.0)]:
        red(0, None, random.randint(40, 95), _dt(2026, 5, random.randint(2, 28)),
            "", vrsta, izn, 0, izn, "virman", djel=random.choice(DJELATNICI))

    BASE.mkdir(parents=True, exist_ok=True)
    wb.save(BASE / "URA_demo.xlsx")
    return r - 3


def tablica_tereni():
    """tereni_demo.xlsx — list 'tereni 2026' (zaglavlje red 1, podaci od reda 2)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tereni 2026"
    ws.append(["#", "polazak", "povratak", "djelatnik", "auto",
               "km prije", "km poslije", "lokacija", "razlog"])
    auti = ["kombi1", "osobno1", "dostavno", "kombi2"]
    lokacije = ["Zagreb - Sljeme", "Karlovac", "Sisak", "Velika Gorica", "Samobor"]
    razlozi = ["obilazak gradilišta", "sastanak s investitorom",
               "ispitivanje materijala", "obilazak gradilišta", "kontrolno mjerenje"]
    for i in range(5):
        d = _dt(2026, 6, 3 + i * 3)
        ws.append([i + 1, d, d, random.choice(DJELATNICI), auti[i % len(auti)],
                   100000 + i * 500, 100000 + i * 500 + random.randint(60, 300),
                   lokacije[i], razlozi[i]])
    wb.save(BASE / "tereni_demo.xlsx")
    return 5


def parra_json():
    """parra_demo.json — lažni ulazni e-računi (zamjena za Parra API u DEMO načinu)."""
    racuni = []
    for i in range(8):
        dob, oib = random.choice(DOBAVLJACI)
        neto = round(random.uniform(50, 2000), 2)
        pdv = round(neto * 0.25, 2)
        d = date(2026, 6, random.randint(2, 27))
        racuni.append({
            "id": f"demo-{i+1:03d}",
            "invoiceNumber": f"{random.randint(100, 9999)}/1/2026",
            "documentDateInCet": d.isoformat() + "T00:00:00",
            "dueDate": (d + timedelta(days=15)).isoformat(),
            "supplierRegistrationName": dob,
            "supplierOib": oib,
            "totalAmountWithoutTax": neto,
            "totalTaxAmount": pdv,
            "totalAmount": round(neto + pdv, 2),
            "status": "NEW",
            "schemaType": "eRacun",
        })
    json.dump(racuni, open(BASE / "parra_demo.json", "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)
    return len(racuni)


if __name__ == "__main__":
    stvori_mape()
    n_ura = knjiga_ura()
    n_ter = tablica_tereni()
    n_par = parra_json()
    print("Demo podaci stvoreni u:", BASE)
    print(f"  URA_demo.xlsx     — {n_ura} redaka")
    print(f"  tereni_demo.xlsx  — {n_ter} terena (lipanj 2026)")
    print(f"  parra_demo.json   — {n_par} lažnih e-računa")
    print("Pokreni aplikaciju (DEMO=1 je zadano) i sve radi na ovim podacima.")
